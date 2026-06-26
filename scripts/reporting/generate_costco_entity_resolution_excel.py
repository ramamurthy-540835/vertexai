#!/usr/bin/env python3
"""Generate a Costco Lead-to-POS Entity Resolution analysis workbook.

Reads fuzzy file-mode outputs (output CSV, debug candidates CSV, summary JSON)
and business rules JSON, then produces a styled four-sheet Excel workbook:

  1. Read Me          — business-friendly explanation
  2. Tuning Parameters — rules JSON reference
  3. Scoring Detail    — one row per lead-POS candidate per scoring set
  4. Pair Analysis     — one row per final lead-POS pair

Does NOT change matching logic. Does NOT write to Cloud SQL or GCS.

Usage:
    python3 scripts/generate_costco_entity_resolution_excel.py \\
      --pair-output-csv  mock_data/115_from_exact/fuzzy_file_mode_output.csv \\
      --debug-candidates-csv mock_data/115_from_exact/fuzzy_file_mode_debug_candidates.csv \\
      --summary-json     mock_data/115_from_exact/fuzzy_file_mode_summary.json \\
      --rules-json       lead_match_runtime/lead_to_pos_match_rules.json \\
      --warehouse-number 115 \\
      --output-xlsx      reports/parallel_run_analysis/115/Entity_Resolution_Pair_Analysis_Costco.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════
# Constants
# ═══════════════════════════════════════════════════════════════

MATCHING_SETS = [
    {"set": 1, "source": "POS native",
     "lead_name": "business_name", "pos_name": "business_name",
     "lead_addr": "full_address",  "pos_addr": "full_address",
     "email": "email",       "phone": "phone"},
    {"set": 2, "source": "POS + OMS company",
     "lead_name": "business_name", "pos_name": "oms_company",
     "lead_addr": "full_address",  "pos_addr": "full_address",
     "email": "email",       "phone": "phone"},
    {"set": 3, "source": "OMS primary business",
     "lead_name": "business_name", "pos_name": "business_name",
     "lead_addr": "full_address",  "pos_addr": "full_oms_address",
     "email": "oms_email_1", "phone": "oms_phone_1"},
    {"set": 4, "source": "OMS primary company",
     "lead_name": "business_name", "pos_name": "oms_company",
     "lead_addr": "full_address",  "pos_addr": "full_oms_address",
     "email": "oms_email_1", "phone": "oms_phone_1"},
    {"set": 5, "source": "OMS secondary business",
     "lead_name": "business_name", "pos_name": "business_name",
     "lead_addr": "full_address",  "pos_addr": "full_oms2_address",
     "email": "oms_email_2", "phone": "oms_phone_2"},
    {"set": 6, "source": "OMS secondary company",
     "lead_name": "business_name", "pos_name": "oms2_company",
     "lead_addr": "full_address",  "pos_addr": "full_oms2_address",
     "email": "oms_email_2", "phone": "oms_phone_2"},
]

SIX_SET_COLS = [f"set_{s}_{c}" for s in range(1, 7)
                for c in ("name_score", "address_score", "base_score",
                          "email_boost", "phone_boost", "final_score")]

REQUIRED_OUTPUT_COLS = [
    "lead_id", "pos_id", "match_result", "similarity_score", "winning_set",
    "match_type", "primary_transaction", "warehouse_number",
]

REQUIRED_DEBUG_COLS = [
    "lead_id", "pos_id", "warehouse_number", "combined_similarity",
    "name_score", "address_score", "email_boost", "phone_boost",
    "final_score", "winning_set", "decision", "reason",
]

# ═══════════════════════════════════════════════════════════════
# Styles
# ═══════════════════════════════════════════════════════════════

FILL_GREEN    = PatternFill("solid", fgColor="C6EFCE")
FILL_AMBER    = PatternFill("solid", fgColor="FFEB9C")
FILL_RED      = PatternFill("solid", fgColor="FFC7CE")
FILL_BLUE     = PatternFill("solid", fgColor="BDD7EE")
FILL_GRAY     = PatternFill("solid", fgColor="D9D9D9")
FILL_HDR      = PatternFill("solid", fgColor="4472C4")
FILL_TITLE    = PatternFill("solid", fgColor="002060")
FILL_LABEL    = PatternFill("solid", fgColor="D6E4F0")
FILL_PARAM    = PatternFill("solid", fgColor="E2EFDA")
FILL_WHAT_IF  = PatternFill("solid", fgColor="BDD7EE")
FILL_WARN     = PatternFill("solid", fgColor="FFF2CC")

FONT_TITLE     = Font(bold=True, size=14, color="FFFFFF")
FONT_SUBTITLE  = Font(bold=True, size=11, color="333333")
FONT_HEADER    = Font(bold=True, size=10, color="FFFFFF")
FONT_BODY      = Font(size=10, color="333333")
FONT_BOLD      = Font(bold=True, size=10, color="333333")
FONT_LINK      = Font(size=10, color="0563C1", underline="single")
FONT_WARN      = Font(bold=True, size=10, color="9C5700")

ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_CTR  = Alignment(horizontal="center", vertical="center")
ALIGN_NUM  = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)


# ═══════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Generate Costco Entity Resolution analysis workbook.")
    p.add_argument("--pair-output-csv", required=True,
                   help="fuzzy_file_mode_output.csv from fuzzy_file_runner.py")
    p.add_argument("--debug-candidates-csv", required=True,
                   help="fuzzy_file_mode_debug_candidates.csv from fuzzy_file_runner.py")
    p.add_argument("--summary-json", required=True,
                   help="fuzzy_file_mode_summary.json from fuzzy_file_runner.py")
    p.add_argument("--rules-json", required=True,
                   help="lead_to_pos_match_rules.json (business rules)")
    p.add_argument("--warehouse-number", type=int, default=None,
                   help="Warehouse number (for per-warehouse output)")
    p.add_argument("--output-xlsx", required=True,
                   help="Output Excel workbook path")
    return p.parse_args()


# ═══════════════════════════════════════════════════════════════
# Data loading and validation
# ═══════════════════════════════════════════════════════════════

def load_rules(path: str | None = None) -> dict:
    from lead_match_runtime.business_rules import load_business_rules
    return load_business_rules(path)


def load_summary(path: str) -> dict:
    with open(path) as f:
        return json.load(f)


def load_pair_output(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str)
    df.columns = df.columns.str.strip().str.lower()
    for col in ("similarity_score",):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def load_debug_candidates(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str)
    df.columns = df.columns.str.strip().str.lower()
    numeric_cols = (["combined_similarity", "name_score", "address_score",
                     "email_boost", "phone_boost", "final_score"]
                    + [c for c in df.columns if c.startswith("set_")])
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


class ValidationError(Exception):
    pass


def validate_data(pair_df: pd.DataFrame, debug_df: pd.DataFrame,
                  summary: dict) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []

    missing_out = [c for c in REQUIRED_OUTPUT_COLS if c not in pair_df.columns]
    if missing_out:
        errors.append(f"Missing required output columns: {missing_out}")

    missing_dbg = [c for c in REQUIRED_DEBUG_COLS if c not in debug_df.columns]
    if missing_dbg:
        errors.append(f"Missing required debug columns: {missing_dbg}")

    if "similarity_score" in pair_df.columns:
        max_score = pair_df["similarity_score"].max()
        if pd.notna(max_score) and max_score >= 100:
            errors.append(f"Fuzzy score >= 100 detected: {max_score}")

    if "warehouse_number" in pair_df.columns:
        if pair_df["warehouse_number"].isna().any() or (pair_df["warehouse_number"] == "").any():
            errors.append("warehouse_number is missing in some pair output rows")

    if "expected_relation" not in debug_df.columns or debug_df["expected_relation"].isna().all():
        warnings.append("expected_relation column missing or entirely empty in debug data")

    has_six_set = any(c in debug_df.columns for c in SIX_SET_COLS)
    if not has_six_set:
        warnings.append("Six-set detail columns not found in debug file")

    oms_cols = [c for c in debug_df.columns
                if "oms" in c.lower() and c not in REQUIRED_DEBUG_COLS]
    if not oms_cols and not has_six_set:
        warnings.append("No OMS variant columns detected")

    return errors, warnings


# ═══════════════════════════════════════════════════════════════
# Sheet 1: Read Me
# ═══════════════════════════════════════════════════════════════

def _set_col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def build_read_me(wb: Workbook, rules: dict, summary: dict,
                  warnings: list[str], has_six_set: bool,
                  warehouse: int | None) -> None:
    ws = wb.active
    ws.title = "Read Me"

    _set_col_width(ws, 1, 4)
    _set_col_width(ws, 2, 90)
    _set_col_width(ws, 3, 4)

    row = 1
    wh_label = f"Warehouse {warehouse}" if warehouse else "All Warehouses"
    title = f"ENTITY RESOLUTION — COSTCO LEAD-TO-POS PAIR ANALYSIS  |  {wh_label}"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = FONT_TITLE
    cell.fill = FILL_TITLE
    ws.cell(row=row, column=2).fill = FILL_TITLE

    row += 1
    subtitle = ("Vertex AI semantic embeddings + deterministic phone/email boosters  "
                "|  Six-set scoring  |  Analysis & explainability only")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws.cell(row=row, column=1, value=subtitle)
    cell.font = Font(italic=True, size=10, color="666666")

    row += 2
    sections = [
        ("What this workbook does", [
            "This workbook analyzes candidate Lead-to-POS pairs produced by the Costco "
            "fuzzy/semantic matching pipeline.",
            "It is for analysis and explainability only — it does NOT write to Cloud SQL, "
            "does NOT upload to GCS, and does NOT change production matching behavior.",
        ]),
        ("How matching works", [
            "1. Exact matching runs first (deterministic field equality) and owns score 100 "
            "(auto-close, Closed - Match).",
            "2. Fuzzy/semantic matching runs on the residual (unmatched, unprocessed, "
            "post-lead records).",
            "3. Vertex AI embeddings (gemini-embedding-001, 768-dim, L2-normalized) create "
            "vector representations of business names and addresses.",
            "4. A combined embedding (name + address) is used as a recall gate to retrieve "
            "the top-K POS candidates per lead. The combined embedding is NOT part of the "
            "final score.",
            "5. Six matching sets score each lead against different POS name/address variants "
            "(native, OMS primary, OMS secondary). The highest-scoring set wins.",
            "6. Formula: base_score = (4 × address_score + 3 × name_score) / 7",
            "7. Email and phone are NOT embedded. They are deterministic boosters: "
            "+5 each for exact match (cap 99.999).",
            "8. Fuzzy scores 70–99.999 → Potential (manual review). Below 70 → No Match. "
            "Fuzzy NEVER produces score 100.",
        ]),
        ("Fiscal classification (applied before scoring)", [
            "• Closed-Existing (CE): POS is before the lead within 6 fiscal periods "
            "→ stub row, lead removed from active set.",
            "• Out-of-Fiscal-Window (OAF): POS is before the lead by more than 6 periods "
            "→ pair dropped, lead stays active.",
            "• Normal: POS is at or after the lead → proceeds to six-set scoring.",
        ]),
        ("Sheet guide", [
            "• Tuning Parameters — Reference configuration from the rules JSON.",
            "• Scoring Detail — One row per lead-POS candidate per scoring set (six rows "
            "per candidate when six-set data is available).",
            "• Pair Analysis — One row per final lead-POS pair, with classification colors.",
        ]),
        ("Color conventions", [
            "• Green — Match / Exact (score 100, exact engine only)",
            "• Amber — Potential / Fuzzy (score 70–99.999)",
            "• Blue  — Closed-Existing (CE)",
            "• Red   — No Match / rejected (below 70)",
            "• Gray  — OAF / dropped",
        ]),
    ]

    for heading, lines in sections:
        cell = ws.cell(row=row, column=2, value=heading)
        cell.font = FONT_SUBTITLE
        row += 1
        for line in lines:
            cell = ws.cell(row=row, column=2, value=line)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_WRAP
            row += 1
        row += 1

    if warnings:
        cell = ws.cell(row=row, column=2, value="Warnings")
        cell.font = FONT_WARN
        cell.fill = FILL_WARN
        row += 1
        for w in warnings:
            cell = ws.cell(row=row, column=2, value=f"⚠ {w}")
            cell.font = FONT_WARN
            cell.fill = FILL_WARN
            row += 1
        row += 1

    if not has_six_set:
        cell = ws.cell(row=row, column=2,
                       value="Six-set detail was not fully available in the input debug file.")
        cell.font = FONT_WARN
        cell.fill = FILL_WARN
        row += 1

    cell = ws.cell(row=row, column=2,
                   value=f"Generated: {summary.get('generated_at', 'N/A')}  |  "
                         f"Model: {summary.get('embedding_model', 'N/A')}  |  "
                         f"Rules schema: {rules.get('schema_version', 'N/A')}")
    cell.font = Font(size=9, color="999999", italic=True)


# ═══════════════════════════════════════════════════════════════
# Sheet 2: Tuning Parameters
# ═══════════════════════════════════════════════════════════════

def build_tuning_parameters(wb: Workbook, rules: dict) -> None:
    ws = wb.create_sheet("Tuning Parameters")

    _set_col_width(ws, 1, 4)
    _set_col_width(ws, 2, 45)
    _set_col_width(ws, 3, 55)
    _set_col_width(ws, 4, 4)

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=1, value="TUNING PARAMETERS")
    cell.font = FONT_TITLE
    cell.fill = FILL_TITLE
    for c in range(2, 4):
        ws.cell(row=row, column=c).fill = FILL_TITLE
    row += 1

    cell = ws.cell(row=row, column=2,
                   value="Reference configuration from rules JSON — "
                         "do not modify for production without approval.")
    cell.font = Font(italic=True, size=10, color="666666")
    row += 2

    embeddings = rules.get("embeddings", {})
    candidate = rules.get("candidate_retrieval", {})
    scoring = rules.get("scoring", {})
    decision = rules.get("decision_rules", {})
    fiscal = rules.get("fiscal_rules", {})
    markers = rules.get("matching_markers", {})

    params = [
        ("Embedding Configuration", None),
        ("Embedding model", embeddings.get("model", "N/A"), False),
        ("Output dimensionality", embeddings.get("output_dimensionality", "N/A"), False),
        ("Task type", embeddings.get("task_type", "N/A"), False),
        ("L2 normalize", embeddings.get("l2_normalize", "N/A"), False),
        ("Model lock", "LOCKED — cross-model cosine is invalid", False),
        ("", None),
        ("Candidate Retrieval", None),
        ("Recall gate field", candidate.get("recall_gate_field", "N/A"), False),
        ("Recall gate minimum similarity", candidate.get("recall_gate_min_similarity", "N/A"), True),
        ("Nearest neighbor limit (top_k)", candidate.get("nearest_neighbor_limit", "N/A"), True),
        ("Distance metric", candidate.get("method", "N/A"), False),
        ("", None),
        ("Scoring Formula", None),
        ("Formula", scoring.get("precision_score_formula", "N/A"), False),
        ("Name weight", embeddings.get("fields", {}).get("name_variant", {}).get("weight", 3), False),
        ("Address weight", embeddings.get("fields", {}).get("address_variant", {}).get("weight", 4), False),
        ("Address contribution", "57.1%", False),
        ("Name contribution", "42.9%", False),
        ("", None),
        ("Deterministic Boosts", None),
        ("Email exact match boost",
         scoring.get("deterministic_boosts", {}).get("email_exact_match", 5), True),
        ("Phone exact match boost",
         scoring.get("deterministic_boosts", {}).get("phone_exact_match", 5), True),
        ("Boost cap", scoring.get("deterministic_boosts", {}).get("cap", 99.999), False),
        ("Disagreement penalty", "None (neutral)", False),
        ("", None),
        ("Decision Thresholds", None),
        ("Fuzzy qualify min score", decision.get("fuzzy_qualify_min_score", 70), True),
        ("Fuzzy max score", decision.get("fuzzy_max_score", 99.999), False),
        ("Exact score (exact engine only)", decision.get("exact_score", 100), False),
        ("No match max score", decision.get("no_match_max_score", 69.999), False),
        ("", None),
        ("Fiscal Rules", None),
        ("Periods per year", fiscal.get("periods_per_year", 13), False),
        ("CE period window", fiscal.get("ce_period_window", 6), True),
        ("Period gap formula", fiscal.get("period_gap_formula", "N/A"), False),
        ("", None),
        ("Do Not Embed", None),
        ("Fields excluded from embedding",
         ", ".join(markers.get("do_not_embed", [])), False),
    ]

    for entry in params:
        if len(entry) == 2:
            label, val = entry
            if val is None:
                cell = ws.cell(row=row, column=2, value=label)
                cell.font = FONT_SUBTITLE
                cell.fill = FILL_LABEL
                ws.cell(row=row, column=3).fill = FILL_LABEL
                row += 1
                continue
            if label == "":
                row += 1
                continue
        else:
            label, val, tunable = entry
            lc = ws.cell(row=row, column=2, value=label)
            lc.font = FONT_BODY
            lc.border = THIN_BORDER

            vc = ws.cell(row=row, column=3, value=val)
            vc.font = FONT_BODY
            vc.border = THIN_BORDER
            if tunable:
                vc.fill = FILL_WHAT_IF
            else:
                vc.fill = FILL_PARAM

            row += 1

    row += 1
    sets_header = ws.cell(row=row, column=2, value="Matching Sets (six-set architecture)")
    sets_header.font = FONT_SUBTITLE
    sets_header.fill = FILL_LABEL
    ws.cell(row=row, column=3).fill = FILL_LABEL
    row += 1

    set_hdrs = ["Set", "Source", "Name Field", "Address Field", "Email", "Phone"]
    for ci, h in enumerate(set_hdrs, 2):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HDR
        c.border = THIN_BORDER
    row += 1

    rule_sets = rules.get("matching_sets", {}).get("sets", [])
    for s in rule_sets:
        vals = [s.get("set"), s.get("source"), s.get("name_field"),
                s.get("address_field"), s.get("email_field"), s.get("phone_field")]
        for ci, v in enumerate(vals, 2):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = FONT_BODY
            c.border = THIN_BORDER
        row += 1


# ═══════════════════════════════════════════════════════════════
# Sheet 3: Scoring Detail
# ═══════════════════════════════════════════════════════════════

SCORING_DETAIL_COLS = [
    "lead_id", "pos_id", "warehouse_number", "candidate_rank",
    "expected_relation", "set_number", "set_source",
    "lead_name_source", "pos_name_source",
    "lead_address_source", "pos_address_source",
    "name_score", "address_score", "base_score",
    "email_source", "phone_source",
    "email_boost", "phone_boost", "final_score",
    "winning_set_flag", "decision", "reason",
]


def _expand_six_set_rows(debug_df: pd.DataFrame, has_six_set: bool) -> pd.DataFrame:
    rows = []
    candidates = debug_df[debug_df["decision"].isin(["Potential", "No Match", "NO_EMBEDDING"])]

    for rank, (_, r) in enumerate(candidates.iterrows(), 1):
        lead_id = r.get("lead_id", "")
        pos_id = r.get("pos_id", "")
        wh = r.get("warehouse_number", "")
        expected = r.get("expected_relation", "")
        winning = r.get("winning_set")
        decision = r.get("decision", "")
        reason = r.get("reason", "")

        if has_six_set:
            for sn in range(1, 7):
                ns = r.get(f"set_{sn}_name_score")
                ads = r.get(f"set_{sn}_address_score")
                bs = r.get(f"set_{sn}_base_score")
                eb = r.get(f"set_{sn}_email_boost", 0)
                pb = r.get(f"set_{sn}_phone_boost", 0)
                fs = r.get(f"set_{sn}_final_score")

                if pd.isna(ns) and pd.isna(ads):
                    continue

                set_def = MATCHING_SETS[sn - 1]
                is_winner = (str(sn) == str(winning)) if pd.notna(winning) else False

                rows.append({
                    "lead_id": lead_id,
                    "pos_id": pos_id,
                    "warehouse_number": wh,
                    "candidate_rank": rank,
                    "expected_relation": expected,
                    "set_number": sn,
                    "set_source": set_def["source"],
                    "lead_name_source": set_def["lead_name"],
                    "pos_name_source": set_def["pos_name"],
                    "lead_address_source": set_def["lead_addr"],
                    "pos_address_source": set_def["pos_addr"],
                    "name_score": ns if pd.notna(ns) else "",
                    "address_score": ads if pd.notna(ads) else "",
                    "base_score": bs if pd.notna(bs) else "",
                    "email_source": set_def["email"],
                    "phone_source": set_def["phone"],
                    "email_boost": eb if pd.notna(eb) else 0,
                    "phone_boost": pb if pd.notna(pb) else 0,
                    "final_score": fs if pd.notna(fs) else "",
                    "winning_set_flag": is_winner,
                    "decision": decision if is_winner else "",
                    "reason": reason if is_winner else "",
                })
        else:
            ns = r.get("name_score")
            ads = r.get("address_score")
            eb = r.get("email_boost", 0)
            pb = r.get("phone_boost", 0)
            fs = r.get("final_score")
            ws_num = int(float(winning)) if pd.notna(winning) and str(winning).strip() else 0
            set_def = MATCHING_SETS[ws_num - 1] if 1 <= ws_num <= 6 else MATCHING_SETS[0]

            rows.append({
                "lead_id": lead_id,
                "pos_id": pos_id,
                "warehouse_number": wh,
                "candidate_rank": rank,
                "expected_relation": expected,
                "set_number": ws_num or "",
                "set_source": set_def["source"] if ws_num else "",
                "lead_name_source": set_def["lead_name"] if ws_num else "",
                "pos_name_source": set_def["pos_name"] if ws_num else "",
                "lead_address_source": set_def["lead_addr"] if ws_num else "",
                "pos_address_source": set_def["pos_addr"] if ws_num else "",
                "name_score": ns if pd.notna(ns) else "",
                "address_score": ads if pd.notna(ads) else "",
                "base_score": fs if pd.notna(fs) else "",
                "email_source": "",
                "phone_source": "",
                "email_boost": eb if pd.notna(eb) else 0,
                "phone_boost": pb if pd.notna(pb) else 0,
                "final_score": fs if pd.notna(fs) else "",
                "winning_set_flag": True,
                "decision": decision,
                "reason": reason,
            })

    ce_oaf = debug_df[debug_df["decision"].isin(["CE", "OAF"])]
    for _, r in ce_oaf.iterrows():
        rows.append({
            "lead_id": r.get("lead_id", ""),
            "pos_id": r.get("pos_id", ""),
            "warehouse_number": r.get("warehouse_number", ""),
            "candidate_rank": "",
            "expected_relation": r.get("expected_relation", ""),
            "set_number": "",
            "set_source": "",
            "lead_name_source": "",
            "pos_name_source": "",
            "lead_address_source": "",
            "pos_address_source": "",
            "name_score": "",
            "address_score": "",
            "base_score": "",
            "email_source": "",
            "phone_source": "",
            "email_boost": "",
            "phone_boost": "",
            "final_score": "",
            "winning_set_flag": "",
            "decision": r.get("decision", ""),
            "reason": r.get("reason", ""),
        })

    return pd.DataFrame(rows, columns=SCORING_DETAIL_COLS)


def _score_fill(decision: str) -> PatternFill | None:
    d = str(decision).strip().lower()
    if d in ("match", "exact"):
        return FILL_GREEN
    if d == "potential":
        return FILL_AMBER
    if d in ("no match", "no_match"):
        return FILL_RED
    if d == "ce":
        return FILL_BLUE
    if d == "oaf":
        return FILL_GRAY
    return None


def build_scoring_detail(wb: Workbook, debug_df: pd.DataFrame,
                         has_six_set: bool) -> int:
    ws = wb.create_sheet("Scoring Detail")

    detail_df = _expand_six_set_rows(debug_df, has_six_set)

    row = 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=len(SCORING_DETAIL_COLS))
    cell = ws.cell(row=row, column=1,
                   value="SCORING DETAIL — PER-SET EVIDENCE")
    cell.font = FONT_TITLE
    cell.fill = FILL_TITLE
    for ci in range(2, len(SCORING_DETAIL_COLS) + 1):
        ws.cell(row=row, column=ci).fill = FILL_TITLE

    row += 1
    note = ("One row per lead-POS candidate per scoring set. "
            "Scores formatted to 3 decimals. "
            "winning_set_flag = TRUE for the selected best set.")
    if not has_six_set:
        note += "  ⚠ Only winning-set data available."
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=len(SCORING_DETAIL_COLS))
    cell = ws.cell(row=row, column=1, value=note)
    cell.font = Font(italic=True, size=9, color="666666")

    row += 1  # blank
    row += 1

    score_num_cols = {"name_score", "address_score", "base_score",
                      "email_boost", "phone_boost", "final_score"}
    wide_cols = {"lead_id": 22, "pos_id": 22, "reason": 40,
                 "set_source": 22, "lead_name_source": 18, "pos_name_source": 18,
                 "lead_address_source": 18, "pos_address_source": 18,
                 "expected_relation": 16, "decision": 14}

    for ci, col_name in enumerate(SCORING_DETAIL_COLS, 1):
        c = ws.cell(row=row, column=ci, value=col_name)
        c.font = FONT_HEADER
        c.fill = FILL_HDR
        c.border = THIN_BORDER
        c.alignment = ALIGN_CTR
        w = wide_cols.get(col_name, 14)
        _set_col_width(ws, ci, w)

    header_row = row
    row += 1

    for _, dr in detail_df.iterrows():
        fill = _score_fill(dr.get("decision", ""))
        for ci, col_name in enumerate(SCORING_DETAIL_COLS, 1):
            val = dr.get(col_name, "")
            if col_name in score_num_cols and val != "" and pd.notna(val):
                try:
                    val = round(float(val), 3)
                except (ValueError, TypeError):
                    pass
            if col_name == "winning_set_flag" and val is True:
                val = "TRUE"
            elif col_name == "winning_set_flag" and val is not True:
                val = "" if val == "" else "FALSE"

            c = ws.cell(row=row, column=ci, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            if col_name in score_num_cols and isinstance(val, (int, float)):
                c.number_format = "0.000"
                c.alignment = ALIGN_NUM
            if fill and col_name == "decision":
                c.fill = fill
        row += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(SCORING_DETAIL_COLS))}{row - 1}"

    return len(detail_df)


# ═══════════════════════════════════════════════════════════════
# Sheet 4: Pair Analysis
# ═══════════════════════════════════════════════════════════════

PAIR_COLS = [
    "lead_id", "pos_id", "warehouse_number",
    "lead_business_name", "pos_business_name",
    "lead_address", "pos_address",
    "combined_similarity", "winning_set",
    "name_score", "address_score",
    "email_boost", "phone_boost",
    "similarity_score", "match_result", "match_type",
    "lifecycle_state", "primary_transaction",
    "closed_existing_flag", "expected_relation", "analyst_note",
]


def _lifecycle_state(match_result: str, match_type: str) -> str:
    mr = str(match_result).strip().lower()
    mt = str(match_type).strip().lower()
    if mr == "match" and mt == "exact":
        return "Closed - Match"
    if mr == "potential":
        return "Potential"
    if mr == "no match":
        return "No Match"
    return ""


def _pair_fill(match_result: str, closed_existing: str) -> PatternFill | None:
    mr = str(match_result).strip().lower()
    ce = str(closed_existing).strip().lower()
    if ce in ("true", "1", "yes"):
        return FILL_BLUE
    if mr == "match":
        return FILL_GREEN
    if mr == "potential":
        return FILL_AMBER
    if mr in ("no match", "no_match"):
        return FILL_RED
    return None


def build_pair_analysis(wb: Workbook, pair_df: pd.DataFrame,
                        debug_df: pd.DataFrame, summary: dict) -> int:
    ws = wb.create_sheet("Pair Analysis")

    debug_lookup = {}
    for _, r in debug_df.iterrows():
        key = (str(r.get("lead_id", "")), str(r.get("pos_id", "")))
        debug_lookup[key] = r

    pair_rows = []
    for _, pr in pair_df.iterrows():
        lead_id = str(pr.get("lead_id", ""))
        pos_id = str(pr.get("pos_id", ""))
        match_result = str(pr.get("match_result", ""))
        match_type = str(pr.get("match_type", ""))
        ce_flag = str(pr.get("closed_existing_flag", ""))

        dr = debug_lookup.get((lead_id, pos_id))

        pair_rows.append({
            "lead_id": lead_id,
            "pos_id": pos_id,
            "warehouse_number": pr.get("warehouse_number", ""),
            "lead_business_name": "",
            "pos_business_name": pr.get("business_name_transaction", ""),
            "lead_address": "",
            "pos_address": ", ".join(filter(None, [
                str(pr.get("address_line_one", "")),
                str(pr.get("city", "")),
                str(pr.get("state", "")),
                str(pr.get("zip_code", "")),
            ])),
            "combined_similarity": dr["combined_similarity"] if dr is not None and pd.notna(dr.get("combined_similarity")) else "",
            "winning_set": pr.get("winning_set", ""),
            "name_score": dr["name_score"] if dr is not None and pd.notna(dr.get("name_score")) else "",
            "address_score": dr["address_score"] if dr is not None and pd.notna(dr.get("address_score")) else "",
            "email_boost": dr["email_boost"] if dr is not None and pd.notna(dr.get("email_boost")) else 0,
            "phone_boost": dr["phone_boost"] if dr is not None and pd.notna(dr.get("phone_boost")) else 0,
            "similarity_score": pr.get("similarity_score", ""),
            "match_result": match_result,
            "match_type": match_type,
            "lifecycle_state": _lifecycle_state(match_result, match_type),
            "primary_transaction": pr.get("primary_transaction", ""),
            "closed_existing_flag": ce_flag,
            "expected_relation": dr["expected_relation"] if dr is not None else "",
            "analyst_note": "",
        })

    result_df = pd.DataFrame(pair_rows, columns=PAIR_COLS)

    row = 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=len(PAIR_COLS))
    cell = ws.cell(row=row, column=1,
                   value="PAIR ANALYSIS — MATCH SCORES & CLASSIFICATION")
    cell.font = FONT_TITLE
    cell.fill = FILL_TITLE
    for ci in range(2, len(PAIR_COLS) + 1):
        ws.cell(row=row, column=ci).fill = FILL_TITLE

    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=len(PAIR_COLS))
    cell = ws.cell(row=row, column=1,
                   value="One row per final lead-POS pair. "
                         "Classification driven by tunable thresholds from rules JSON.")
    cell.font = Font(italic=True, size=9, color="666666")

    row += 1  # blank
    row += 1

    # Portfolio summary above the table
    stats_data = [
        ("Total pairs evaluated", summary.get("candidate_pairs",
                                               summary.get("fuzzy_candidates_evaluated", "N/A"))),
        ("Exact-owned excluded", summary.get("exact_excluded_pos",
                                              summary.get("exact_owned_pos_excluded", "N/A"))),
        ("Fuzzy candidates scored", summary.get("normal_scored",
                                                 summary.get("fuzzy_candidates_evaluated", "N/A"))),
        ("Potential matches", summary.get("above_threshold",
                                          summary.get("potential_matches", "N/A"))),
        ("Closed-Existing stubs", summary.get("ce_stubs",
                                               summary.get("closed_existing_rows", 0))),
        ("OAF dropped", summary.get("oaf_dropped", 0)),
        ("No Match / rejected", summary.get("below_threshold", "N/A")),
        ("Max fuzzy score observed", summary.get("max_score",
                                                  summary.get("fuzzy_max_score_observed", "N/A"))),
        ("Fuzzy reached 100", summary.get("fuzzy_reached_100",
                                           "true" if float(summary.get("max_score", 0) or 0) >= 100
                                           else "false")),
        ("Cloud SQL touched", summary.get("cloud_sql_touched", "false")),
        ("GCS upload", summary.get("gcs_upload", "false")),
    ]

    cell = ws.cell(row=row, column=2, value="Portfolio Summary")
    cell.font = FONT_SUBTITLE
    cell.fill = FILL_LABEL
    ws.cell(row=row, column=3).fill = FILL_LABEL
    row += 1

    for label, val in stats_data:
        lc = ws.cell(row=row, column=2, value=label)
        lc.font = FONT_BODY
        lc.border = THIN_BORDER
        vc = ws.cell(row=row, column=3, value=val)
        vc.font = FONT_BOLD
        vc.border = THIN_BORDER
        row += 1

    row += 2

    score_num_cols = {"combined_similarity", "name_score", "address_score",
                      "email_boost", "phone_boost", "similarity_score"}
    wide_cols = {"lead_id": 22, "pos_id": 22, "lead_business_name": 28,
                 "pos_business_name": 28, "lead_address": 30, "pos_address": 30,
                 "match_result": 14, "match_type": 14, "lifecycle_state": 18,
                 "expected_relation": 16, "analyst_note": 30}

    for ci, col_name in enumerate(PAIR_COLS, 1):
        c = ws.cell(row=row, column=ci, value=col_name)
        c.font = FONT_HEADER
        c.fill = FILL_HDR
        c.border = THIN_BORDER
        c.alignment = ALIGN_CTR
        w = wide_cols.get(col_name, 14)
        _set_col_width(ws, ci, w)

    header_row = row
    row += 1

    for _, pr in result_df.iterrows():
        fill = _pair_fill(pr.get("match_result", ""), pr.get("closed_existing_flag", ""))
        for ci, col_name in enumerate(PAIR_COLS, 1):
            val = pr.get(col_name, "")
            if col_name in score_num_cols and val != "" and pd.notna(val):
                try:
                    val = round(float(val), 3)
                except (ValueError, TypeError):
                    pass

            c = ws.cell(row=row, column=ci, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            if col_name in score_num_cols and isinstance(val, (int, float)):
                c.number_format = "0.000"
                c.alignment = ALIGN_NUM
            if fill:
                c.fill = fill
        row += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(PAIR_COLS))}{row - 1}"

    return len(result_df)


# ═══════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════

def _check_prerequisites(args: argparse.Namespace) -> list[str]:
    missing = []
    for label, path in [
        ("pair-output-csv", args.pair_output_csv),
        ("debug-candidates-csv", args.debug_candidates_csv),
        ("summary-json", args.summary_json),
        ("rules-json", args.rules_json),
    ]:
        if not Path(path).exists():
            missing.append((label, path))
    return missing


def main() -> int:
    args = parse_args()

    missing = _check_prerequisites(args)
    if missing:
        print("ERROR: Required input files not found:\n")
        for label, path in missing:
            print(f"  --{label}  {path}")
        print("\nTo generate fuzzy file-mode outputs, run:\n")
        print("  python3 lead_match_runtime/fuzzy_file_runner.py \\")
        print("    --source-mode files \\")
        print("    --leads-file mock_data/115/leads_corrected.xlsx \\")
        print("    --pos-file mock_data/115/pos_corrected.xlsx \\")
        print("    --exact-output-csv ./primary_match_output.csv \\")
        print("    --warehouse-number 115 \\")
        print("    --output-dir mock_data/115_from_exact \\")
        print("    --top-k 20 --limit-leads 50")
        return 1

    print("Loading inputs...")
    rules = load_rules(args.rules_json)
    summary = load_summary(args.summary_json)
    pair_df = load_pair_output(args.pair_output_csv)
    debug_df = load_debug_candidates(args.debug_candidates_csv)

    print("Validating data...")
    errors, warnings = validate_data(pair_df, debug_df, summary)
    if errors:
        print("\nVALIDATION ERRORS (workbook not generated):")
        for e in errors:
            print(f"  ✗ {e}")
        return 1
    if warnings:
        print("\nValidation warnings:")
        for w in warnings:
            print(f"  ⚠ {w}")

    has_six_set = any(c in debug_df.columns for c in SIX_SET_COLS)

    print("Building workbook...")
    wb = Workbook()

    build_read_me(wb, rules, summary, warnings, has_six_set, args.warehouse_number)
    build_tuning_parameters(wb, rules)
    scoring_rows = build_scoring_detail(wb, debug_df, has_six_set)
    pair_rows = build_pair_analysis(wb, pair_df, debug_df, summary)

    output_path = Path(args.output_xlsx)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    print(f"\n{'='*60}")
    print(f"Costco Entity Resolution Workbook Generated")
    print(f"{'='*60}")
    print(f"  Output:               {output_path}")
    print(f"  Pair Analysis rows:   {pair_rows}")
    print(f"  Scoring Detail rows:  {scoring_rows}")
    print(f"  Six-set detail:       {'Yes' if has_six_set else 'No (winning set only)'}")
    print(f"  Warehouse:            {args.warehouse_number or 'All'}")
    print(f"  Potential matches:    {summary.get('above_threshold', summary.get('potential_matches', 'N/A'))}")
    print(f"  CE stubs:             {summary.get('ce_stubs', 0)}")
    print(f"  OAF dropped:          {summary.get('oaf_dropped', 0)}")
    print(f"  Max fuzzy score:      {summary.get('max_score', 'N/A')}")
    print(f"  Fuzzy reached 100:    {'YES ⚠' if float(summary.get('max_score', 0) or 0) >= 100 else 'No'}")
    print(f"  Cloud SQL touched:    No")
    print(f"  GCS upload:           No")
    print(f"  Production logic:     Unchanged")
    print(f"{'='*60}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
