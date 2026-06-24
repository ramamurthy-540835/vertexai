#!/usr/bin/env python3
"""Update reports/warehouse_test_run_tracker.xlsx from local mock data and report outputs.

Read-only against all source data — only writes the tracker Excel file.

Data sources (layered, later sources fill gaps left by earlier ones):
  1. mock_data/<warehouse>/          -> Lead/POS Excel row counts
  2. mock_data/warehouse_volume_ranking.csv -> Reference POS volume
  3. reports/lead_match/ctoteam/<warehouse>/<run_id>/summary.json + matches.csv
     -> Full match-run details (rows, scores, types, lifecycle states)
  4. --fetch-gh  -> GitHub Actions workflow jobs via `gh` CLI
     Reads GH_TOKEN from .env.local (--env-file) to authenticate.
     Queries lead_match_semantic_workflow.yml runs, matches jobs by name
     pattern `ctoteam-lead-match-<warehouse>`, extracts status + duration.
  5. --fetch-gcs -> GCS bucket reports via `gsutil`

GitHub Actions connection flow (--fetch-gh):
  1. Load GH_TOKEN from .env.local  (key=GH_TOKEN, plain PAT value)
  2. Call: gh run list --repo <repo> --workflow <workflow> --limit 50
     to get recent workflow run IDs and conclusions.
  3. For each run, call: gh run view <run_id> --repo <repo> --json jobs
     to get the job list.  Each job name follows the pattern
     `ctoteam-lead-match-<warehouse>`.  We regex-match the warehouse
     number from the job name.
  4. For each requested warehouse we keep the latest *successful* job
     (falling back to the most recent job of any conclusion).
  5. The job's startedAt / completedAt timestamps give us Duration.
     The run's databaseId + attempt + warehouse give us the match_run_id
     in the format `github-<databaseId>-<attempt>-<warehouse>`.
"""

import argparse
import csv
import json
import logging
import os
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

TRACKER_COLUMNS = [
    "Warehouse Number",
    "Reference POS Volume",
    "Mock Data Folder Exists",
    "Lead Excel Rows",
    "POS Excel Rows",
    "Latest Match Run ID",
    "Environment",
    "Run Status",
    "Run Date/Time",
    "Duration",
    "Cloud SQL Lead Rows",
    "Cloud SQL POS Rows",
    "Lead Embedding Rows",
    "POS Embedding Rows",
    "Total Match Rows",
    "Exact Rows",
    "Fuzzy Rows",
    "Manual Review Rows",
    "Matching High Rows",
    "Potential Medium Rows",
    "Potential Low Rows",
    "Primary Transactions",
    "Closed Match Rows",
    "Potential Rows",
    "Non-Exact Min Score",
    "Non-Exact Max Score",
    "Non-Exact >=100 Count",
    "Below 70 Count",
    "Report Local Path",
    "GCS Summary URI",
    "GH Run URL",
    "Last Updated UTC",
    "Notes",
]

KNOWN_RUNS = {
    "1581": "github-28023202257-1-1581",
}

STATUS_FILLS = {
    "Completed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Not Run": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Error": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Failed": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Dry Run": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "In Progress": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "Cancelled": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(
        description="Update the warehouse test run tracker Excel file."
    )
    p.add_argument(
        "--warehouses",
        help="Comma-separated warehouse numbers (default: auto-detect from mock_data/)",
    )
    p.add_argument("--fetch-gcs", action="store_true", help="Fetch reports from GCS via gsutil")
    p.add_argument("--fetch-gh", action="store_true",
                    help="Fetch workflow run details from GitHub Actions via gh CLI")
    p.add_argument("--gh-repo", default="ramamurthy-540835/vertexai",
                    help="GitHub repo owner/name")
    p.add_argument("--gh-workflow", default="lead_match_semantic_workflow.yml",
                    help="Workflow filename to query")
    p.add_argument("--env-file", default=".env.local",
                    help="Env file to load GH_TOKEN from (default: .env.local)")
    p.add_argument("--bucket", default="lead-match-ctoteam", help="GCS bucket name")
    p.add_argument("--project", default="ctoteam", help="Project name in report path")
    p.add_argument(
        "--tracker",
        default="reports/warehouse_test_run_tracker.xlsx",
        help="Path to tracker Excel file",
    )
    p.add_argument(
        "--reports-root",
        default="reports/lead_match/ctoteam",
        help="Local reports root directory",
    )
    p.add_argument("--mock-root", default="mock_data", help="Mock data root directory")
    return p.parse_args()


# ---------------------------------------------------------------------------
# Mock data helpers
# ---------------------------------------------------------------------------

def discover_warehouses(mock_root: str) -> list[str]:
    """Find warehouse folders under mock_data/ that are numeric."""
    root = Path(mock_root)
    if not root.is_dir():
        log.warning("Mock data root %s not found", mock_root)
        return []
    return [e.name for e in sorted(root.iterdir()) if e.is_dir() and e.name.isdigit()]


def read_volume_ranking(mock_root: str) -> dict[str, int]:
    """Read warehouse_volume_ranking.csv -> {warehouse_number: reference_pos_volume}."""
    path = Path(mock_root) / "warehouse_volume_ranking.csv"
    if not path.exists():
        log.warning("Volume ranking file not found: %s", path)
        return {}
    mapping: dict[str, int] = {}
    with open(path, newline="") as f:
        for row in csv.DictReader(f):
            wh = row.get("warehouse_number", "").strip()
            vol = row.get("reference_pos_volume", "").strip()
            if wh and vol:
                mapping[wh] = int(vol)
    return mapping


def count_excel_rows(filepath: str) -> int | None:
    """Count data rows in an Excel file (excluding header). Returns None if missing."""
    if not os.path.isfile(filepath):
        return None
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        count = max(ws.max_row - 1, 0) if ws.max_row else 0
        wb.close()
        return count
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Local report helpers
# ---------------------------------------------------------------------------

def find_latest_local_run(reports_root: str, warehouse: str) -> Path | None:
    """Find latest report run folder for a warehouse under local reports root."""
    wh_dir = Path(reports_root) / warehouse
    if not wh_dir.is_dir():
        return None

    known = KNOWN_RUNS.get(warehouse)
    if known:
        known_dir = wh_dir / known
        if known_dir.is_dir() and (known_dir / "summary.json").exists():
            return known_dir

    candidates: list[tuple[str, Path]] = []
    for entry in wh_dir.iterdir():
        if not entry.is_dir():
            continue
        summary = entry / "summary.json"
        if summary.exists():
            try:
                data = json.loads(summary.read_text())
                ts = data.get("generated_at", "") or ""
                candidates.append((ts, entry))
            except (json.JSONDecodeError, OSError):
                candidates.append(("", entry))
        else:
            candidates.append(("", entry))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def parse_summary(summary_path: Path) -> dict:
    """Parse summary.json into a flat dict of tracker-relevant fields."""
    data = json.loads(summary_path.read_text())
    mtc = data.get("match_type_counts", {})
    lsc = data.get("lifecycle_state_counts", {})
    fsb = data.get("fuzzy_score_band", {})
    report_uris = data.get("report_uris", {})
    return {
        "project": data.get("project", ""),
        "warehouse": data.get("warehouse", ""),
        "match_run_id": data.get("match_run_id", ""),
        "dry_run": data.get("dry_run", False),
        "generated_at": data.get("generated_at", ""),
        "duration": data.get("duration", ""),
        "lead_rows": data.get("lead_rows"),
        "pos_rows": data.get("pos_rows"),
        "lead_embedding_rows": data.get("lead_embedding_rows"),
        "pos_embedding_rows": data.get("pos_embedding_rows"),
        "match_rows": data.get("match_rows"),
        "primary_transaction_count": data.get("primary_transaction_count"),
        "exact": mtc.get("Exact", 0),
        "fuzzy": mtc.get("Fuzzy", 0),
        "manual_review": mtc.get("Manual Review", 0),
        "closed_match": lsc.get("Closed - Match", 0),
        "potential": lsc.get("Potential", 0),
        "fuzzy_floor": fsb.get("floor"),
        "fuzzy_ceiling": fsb.get("ceiling"),
        "gcs_summary_uri": report_uris.get("summary_json", ""),
    }


def analyze_matches_csv(matches_path: Path) -> dict:
    """Analyze matches.csv for score distribution of non-exact matches."""
    result = {
        "csv_exact": 0,
        "csv_fuzzy": 0,
        "csv_manual_review": 0,
        "matching_high": 0,
        "potential_medium": 0,
        "potential_low": 0,
        "below_70": 0,
        "non_exact_gte_100": 0,
        "non_exact_min": None,
        "non_exact_max": None,
    }
    if not matches_path.exists():
        return result

    non_exact_scores: list[float] = []
    with open(matches_path, newline="") as f:
        for row in csv.DictReader(f):
            match_type = row.get("match_type", "").strip()
            if match_type == "Exact":
                result["csv_exact"] += 1
                continue
            elif match_type == "Fuzzy":
                result["csv_fuzzy"] += 1
            elif match_type == "Manual Review":
                result["csv_manual_review"] += 1

            try:
                score = float(row.get("final_score", 0))
            except (ValueError, TypeError):
                continue

            non_exact_scores.append(score)
            if score >= 100:
                result["non_exact_gte_100"] += 1
            elif 90 <= score < 100:
                result["matching_high"] += 1
            elif 85 <= score < 90:
                result["potential_medium"] += 1
            elif 70 <= score < 85:
                result["potential_low"] += 1
            else:
                result["below_70"] += 1

    if non_exact_scores:
        result["non_exact_min"] = min(non_exact_scores)
        result["non_exact_max"] = max(non_exact_scores)
    return result


# ---------------------------------------------------------------------------
# GCS helpers
# ---------------------------------------------------------------------------

def gsutil_ls(gcs_path: str) -> list[str]:
    try:
        result = subprocess.run(
            ["gsutil", "ls", gcs_path],
            capture_output=True, text=True, timeout=30,
        )
        if result.returncode != 0:
            return []
        return [l.strip() for l in result.stdout.strip().splitlines() if l.strip()]
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return []


def gsutil_cp(src: str, dst: str) -> bool:
    try:
        result = subprocess.run(
            ["gsutil", "cp", src, dst],
            capture_output=True, text=True, timeout=60,
        )
        return result.returncode == 0
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def fetch_gcs_latest_run(
    bucket: str, project: str, warehouse: str, reports_root: str
) -> tuple[Path | None, str | None, list[str]]:
    """Fetch latest report from GCS. Returns (local_path, gcs_summary_uri, warnings)."""
    warnings: list[str] = []
    gcs_base = f"gs://{bucket}/reports/lead_match/{project}/{warehouse}/"
    run_dirs = gsutil_ls(gcs_base)
    if not run_dirs:
        warnings.append(f"No GCS runs found at {gcs_base}")
        return None, None, warnings

    run_ids = [uri.rstrip("/").split("/")[-1] for uri in run_dirs if uri.rstrip("/").split("/")[-1]]
    if not run_ids:
        warnings.append(f"No valid run folders at {gcs_base}")
        return None, None, warnings

    known = KNOWN_RUNS.get(warehouse)
    latest_run_id = known if (known and known in run_ids) else sorted(run_ids)[-1]

    gcs_run_base = f"{gcs_base}{latest_run_id}/"
    gcs_summary_uri = f"{gcs_run_base}summary.json"
    local_dir = Path(reports_root) / warehouse / latest_run_id
    local_dir.mkdir(parents=True, exist_ok=True)

    for filename in ("summary.json", "matches.csv", "report.md"):
        src = f"{gcs_run_base}{filename}"
        dst = str(local_dir / filename)
        if not os.path.isfile(dst):
            if not gsutil_cp(src, dst) and filename == "summary.json":
                warnings.append(f"Failed to download {src}")
                return None, gcs_summary_uri, warnings

    return local_dir, gcs_summary_uri, warnings


# ---------------------------------------------------------------------------
# GitHub Actions helpers
# ---------------------------------------------------------------------------

def load_env_file(env_file: str) -> dict[str, str]:
    """Parse a .env / .env.local file into a dict.

    Ignores comments (#) and blank lines. Example .env.local:
        GH_TOKEN=github_pat_11BRE5E3Y...
    """
    env: dict[str, str] = {}
    path = Path(env_file)
    if not path.exists():
        return env
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        key, _, value = line.partition("=")
        env[key.strip()] = value.strip()
    return env


def gh_run_json(cmd_args: list[str], gh_env: dict[str, str]) -> list | dict | None:
    """Run a `gh` CLI command, return parsed JSON or None on failure.

    gh_env is merged into the process environment so GH_TOKEN is available.
    """
    env = {**os.environ, **gh_env}
    try:
        result = subprocess.run(
            ["gh"] + cmd_args,
            capture_output=True, text=True, timeout=30, env=env,
        )
        if result.returncode != 0:
            log.warning("gh command failed: %s", result.stderr.strip())
            return None
        return json.loads(result.stdout)
    except (FileNotFoundError, subprocess.TimeoutExpired, json.JSONDecodeError) as e:
        log.warning("gh command error: %s", e)
        return None


def format_duration(started_at: str, completed_at: str) -> str:
    """Compute human-readable duration from two ISO-8601 timestamps."""
    try:
        start = datetime.fromisoformat(started_at.replace("Z", "+00:00"))
        end = datetime.fromisoformat(completed_at.replace("Z", "+00:00"))
        total_secs = int((end - start).total_seconds())
        if total_secs < 0:
            return ""
        mins, secs = divmod(total_secs, 60)
        return f"{mins}m {secs}s" if mins > 0 else f"{secs}s"
    except (ValueError, TypeError):
        return ""


def fetch_gh_workflow_jobs(
    repo: str,
    workflow: str,
    warehouses: list[str],
    gh_env: dict[str, str],
) -> dict[str, dict]:
    """Query GitHub Actions for the latest job per warehouse.

    Connection flow:
      1. `gh run list` with GH_TOKEN from .env.local to list recent runs.
      2. For each run, `gh run view <id> --json jobs` to get job details.
      3. Job names follow `ctoteam-lead-match-<warehouse>` — regex-match
         the warehouse number.
      4. Prefer latest *successful* job; fall back to most-recent of any status.

    Returns {warehouse: {gh_run_id, conclusion, started_at, completed_at,
                         duration, match_run_id, gh_run_url}}.
    """
    runs = gh_run_json(
        ["run", "list", "--repo", repo, "--workflow", workflow,
         "--limit", "50", "--json", "databaseId,conclusion,status"],
        gh_env,
    )
    if not runs:
        log.warning("No workflow runs returned from GitHub")
        return {}

    wh_set = set(warehouses)
    found: dict[str, dict] = {}
    fallback: dict[str, dict] = {}
    job_name_re = re.compile(r"lead-match-(\d+)$")

    for run in runs:
        if found.keys() >= wh_set:
            break
        run_id = run.get("databaseId")
        run_conclusion = run.get("conclusion", "")
        if not run_id:
            continue

        jobs = gh_run_json(
            ["run", "view", str(run_id), "--repo", repo,
             "--json", "jobs", "--jq", ".jobs[]"],
            gh_env,
        )
        if not jobs:
            continue
        if isinstance(jobs, dict):
            jobs = [jobs]

        for job in jobs:
            job_name = job.get("name", "")
            m = job_name_re.search(job_name)
            if not m:
                continue
            wh = m.group(1)
            if wh not in wh_set:
                continue

            started = job.get("startedAt", "")
            completed = job.get("completedAt", "")
            conclusion = job.get("conclusion", "") or run_conclusion
            duration = format_duration(started, completed)
            match_run_id = f"github-{run_id}-1-{wh}"
            gh_run_url = f"https://github.com/{repo}/actions/runs/{run_id}"

            entry = {
                "gh_run_id": run_id,
                "conclusion": conclusion,
                "started_at": started,
                "completed_at": completed,
                "duration": duration,
                "match_run_id": match_run_id,
                "gh_run_url": gh_run_url,
            }

            if conclusion == "success" and wh not in found:
                found[wh] = entry
                log.info("  GH run %s -> warehouse %s (success, %s)", run_id, wh, duration)
            elif wh not in found and wh not in fallback:
                fallback[wh] = entry
                log.info("  GH run %s -> warehouse %s (%s, %s) [fallback]",
                         run_id, wh, conclusion, duration)

    for wh in wh_set:
        if wh not in found and wh in fallback:
            found[wh] = fallback[wh]

    return found


# ---------------------------------------------------------------------------
# Row builder — merges all data sources for one warehouse
# ---------------------------------------------------------------------------

def build_warehouse_row(
    warehouse: str,
    mock_root: str,
    reports_root: str,
    volume_map: dict[str, int],
    fetch_gcs: bool,
    bucket: str,
    project: str,
    gh_job: dict | None = None,
) -> tuple[dict, list[str]]:
    """Build a single warehouse's tracker row. Returns (row_dict, warnings)."""
    warnings: list[str] = []
    row = {col: "" for col in TRACKER_COLUMNS}
    row["Warehouse Number"] = int(warehouse)
    row["Reference POS Volume"] = volume_map.get(warehouse, "")

    # --- Mock data ---
    mock_dir = Path(mock_root) / warehouse
    mock_exists = mock_dir.is_dir()
    row["Mock Data Folder Exists"] = "Yes" if mock_exists else "No"

    if mock_exists:
        lead_count = count_excel_rows(str(mock_dir / "leads_corrected.xlsx"))
        pos_count = count_excel_rows(str(mock_dir / "pos_corrected.xlsx"))
        row["Lead Excel Rows"] = lead_count if lead_count is not None else ""
        row["POS Excel Rows"] = pos_count if pos_count is not None else ""

    # --- GCS reports ---
    run_dir = None
    gcs_summary_uri = ""
    if fetch_gcs:
        log.info("Fetching GCS reports for warehouse %s ...", warehouse)
        gcs_dir, gcs_uri, gcs_warns = fetch_gcs_latest_run(
            bucket, project, warehouse, reports_root,
        )
        warnings.extend(gcs_warns)
        if gcs_dir:
            run_dir = gcs_dir
        if gcs_uri:
            gcs_summary_uri = gcs_uri

    # --- Local reports ---
    if run_dir is None:
        run_dir = find_latest_local_run(reports_root, warehouse)

    # --- Pre-compute GH fields ---
    gh_conclusion_map = {"success": "Completed", "failure": "Failed", "cancelled": "Cancelled"}
    gh_status = ""
    gh_duration = ""
    gh_started = ""
    gh_run_id_str = ""
    gh_run_url = ""
    if gh_job:
        c = gh_job.get("conclusion", "")
        gh_status = gh_conclusion_map.get(c, c or "In Progress")
        gh_duration = gh_job.get("duration", "")
        gh_started = gh_job.get("started_at", "")
        gh_run_id_str = gh_job.get("match_run_id", "")
        gh_run_url = gh_job.get("gh_run_url", "")

    # --- No local/GCS report found ---
    if run_dir is None:
        if gh_job:
            row["Run Status"] = gh_status
            row["Latest Match Run ID"] = gh_run_id_str
            row["Run Date/Time"] = gh_started
            row["Duration"] = gh_duration
            row["Environment"] = project
            row["GH Run URL"] = gh_run_url
            row["Notes"] = "GH Actions data only; no local/GCS report found"
        else:
            row["Run Status"] = "Not Run"
            row["Notes"] = "; ".join(warnings) if warnings else ""
        row["Last Updated UTC"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        return row, warnings

    # --- Parse summary.json ---
    summary_path = run_dir / "summary.json"
    matches_path = run_dir / "matches.csv"

    if not summary_path.exists():
        row["Run Status"] = "Error"
        row["Report Local Path"] = str(run_dir)
        warnings.append(f"summary.json missing in {run_dir}")
        row["Notes"] = "; ".join(warnings)
        row["Last Updated UTC"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        return row, warnings

    try:
        summary = parse_summary(summary_path)
    except (json.JSONDecodeError, OSError) as e:
        row["Run Status"] = "Error"
        row["Report Local Path"] = str(run_dir)
        warnings.append(f"Failed to parse summary.json: {e}")
        row["Notes"] = "; ".join(warnings)
        row["Last Updated UTC"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        return row, warnings

    # --- Populate from summary ---
    row["Run Status"] = "Dry Run" if summary.get("dry_run") else "Completed"
    row["Latest Match Run ID"] = summary.get("match_run_id", "")
    row["Environment"] = summary.get("project", "")
    row["Run Date/Time"] = summary.get("generated_at", "")
    row["Duration"] = summary.get("duration", "")
    row["Cloud SQL Lead Rows"] = summary.get("lead_rows", "")
    row["Cloud SQL POS Rows"] = summary.get("pos_rows", "")
    row["Lead Embedding Rows"] = summary.get("lead_embedding_rows", "")
    row["POS Embedding Rows"] = summary.get("pos_embedding_rows", "")
    row["Total Match Rows"] = summary.get("match_rows", "")
    row["Primary Transactions"] = summary.get("primary_transaction_count", "")
    row["Closed Match Rows"] = summary.get("closed_match", "")
    row["Potential Rows"] = summary.get("potential", "")
    row["Report Local Path"] = str(run_dir)
    row["GCS Summary URI"] = gcs_summary_uri or summary.get("gcs_summary_uri", "")

    # --- Merge GH data ---
    if gh_job:
        row["GH Run URL"] = gh_run_url
        local_run_id = row["Latest Match Run ID"]
        local_is_gh_run = local_run_id.startswith("github-")

        if local_is_gh_run:
            if not row["Duration"]:
                row["Duration"] = gh_duration
        else:
            if gh_duration:
                notes_extra = f"GH workflow duration: {gh_duration} (run {gh_run_id_str})"
                warnings.append(notes_extra)

    # --- Match type counts from summary ---
    row["Exact Rows"] = summary.get("exact", 0)
    row["Fuzzy Rows"] = summary.get("fuzzy", 0)
    row["Manual Review Rows"] = summary.get("manual_review", 0)

    if summary.get("fuzzy_floor") is not None:
        row["Non-Exact Min Score"] = summary["fuzzy_floor"]
    if summary.get("fuzzy_ceiling") is not None:
        row["Non-Exact Max Score"] = summary["fuzzy_ceiling"]

    # --- Override from matches.csv (more accurate per-row analysis) ---
    if matches_path.exists():
        csv_stats = analyze_matches_csv(matches_path)
        row["Exact Rows"] = csv_stats["csv_exact"]
        row["Fuzzy Rows"] = csv_stats["csv_fuzzy"]
        row["Manual Review Rows"] = csv_stats["csv_manual_review"]
        row["Matching High Rows"] = csv_stats["matching_high"]
        row["Potential Medium Rows"] = csv_stats["potential_medium"]
        row["Potential Low Rows"] = csv_stats["potential_low"]
        row["Below 70 Count"] = csv_stats["below_70"]
        row["Non-Exact >=100 Count"] = csv_stats["non_exact_gte_100"]
        if csv_stats["non_exact_min"] is not None:
            row["Non-Exact Min Score"] = round(csv_stats["non_exact_min"], 4)
        if csv_stats["non_exact_max"] is not None:
            row["Non-Exact Max Score"] = round(csv_stats["non_exact_max"], 4)
    else:
        warnings.append("matches.csv not found; score distribution unavailable")

    row["Notes"] = "; ".join(w for w in warnings if w)
    row["Last Updated UTC"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    return row, warnings


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def write_tracker(tracker_path: str, rows: list[dict]):
    """Write/update the tracker Excel file with formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Warehouse Test Runs"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col_idx, col_name in enumerate(TRACKER_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(TRACKER_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

        status = row_data.get("Run Status", "")
        status_col = TRACKER_COLUMNS.index("Run Status") + 1
        status_cell = ws.cell(row=row_idx, column=status_col)
        if status in STATUS_FILLS:
            status_cell.fill = STATUS_FILLS[status]

    for col_idx in range(1, len(TRACKER_COLUMNS) + 1):
        max_width = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for row_idx in range(2, len(rows) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_width = max(max_width, min(len(val), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_width + 3

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(TRACKER_COLUMNS))}{len(rows) + 1}"

    for sc in ("Non-Exact Min Score", "Non-Exact Max Score"):
        if sc in TRACKER_COLUMNS:
            ci = TRACKER_COLUMNS.index(sc) + 1
            for ri in range(2, len(rows) + 2):
                cell = ws.cell(row=ri, column=ci)
                if isinstance(cell.value, float):
                    cell.number_format = numbers.FORMAT_NUMBER_00

    os.makedirs(os.path.dirname(tracker_path) or ".", exist_ok=True)
    wb.save(tracker_path)
    wb.close()
    log.info("Tracker saved to %s", tracker_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = parse_args()

    if args.warehouses:
        warehouse_list = [w.strip() for w in args.warehouses.split(",") if w.strip()]
    else:
        warehouse_list = discover_warehouses(args.mock_root)

    if not warehouse_list:
        log.error("No warehouses found. Use --warehouses or ensure mock_data/ has warehouse folders.")
        sys.exit(1)

    log.info("Warehouses to process: %s", ", ".join(warehouse_list))
    volume_map = read_volume_ranking(args.mock_root)

    # --- GitHub Actions ---
    gh_jobs: dict[str, dict] = {}
    if args.fetch_gh:
        env_vars = load_env_file(args.env_file)
        gh_env: dict[str, str] = {}
        if "GH_TOKEN" in env_vars:
            gh_env["GH_TOKEN"] = env_vars["GH_TOKEN"]
            log.info("Loaded GH_TOKEN from %s", args.env_file)
        elif os.environ.get("GH_TOKEN"):
            log.info("Using GH_TOKEN from environment")
        else:
            log.warning("No GH_TOKEN found in %s or environment; gh calls may fail", args.env_file)

        log.info("Fetching GitHub Actions workflow jobs from %s ...", args.gh_repo)
        gh_jobs = fetch_gh_workflow_jobs(
            args.gh_repo, args.gh_workflow, warehouse_list, gh_env,
        )
        log.info("Found GH job data for warehouses: %s",
                 ", ".join(sorted(gh_jobs.keys())) or "none")

    # --- Build rows ---
    rows: list[dict] = []
    report_warehouses: list[str] = []
    mock_only_warehouses: list[str] = []
    error_warehouses: list[str] = []

    for wh in warehouse_list:
        log.info("Processing warehouse %s ...", wh)
        try:
            row, warnings = build_warehouse_row(
                warehouse=wh,
                mock_root=args.mock_root,
                reports_root=args.reports_root,
                volume_map=volume_map,
                fetch_gcs=args.fetch_gcs,
                bucket=args.bucket,
                project=args.project,
                gh_job=gh_jobs.get(wh),
            )
            rows.append(row)

            status = row.get("Run Status", "")
            if status in ("Completed", "Dry Run"):
                report_warehouses.append(wh)
            elif status == "Not Run":
                mock_only_warehouses.append(wh)
            else:
                error_warehouses.append(wh)

            for w in warnings:
                log.warning("  [%s] %s", wh, w)

        except Exception as e:
            log.error("  [%s] Unexpected error: %s", wh, e)
            error_warehouses.append(wh)
            rows.append({
                "Warehouse Number": int(wh),
                "Run Status": "Error",
                "Notes": str(e),
                "Last Updated UTC": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            })

    rows.sort(key=lambda r: r.get("Warehouse Number", 0))
    write_tracker(args.tracker, rows)

    # --- Summary ---
    print()
    print("=" * 60)
    print(f"Tracker path:              {args.tracker}")
    print(f"Warehouses updated:        {len(rows)}")
    print(f"With reports:              {', '.join(report_warehouses) or 'none'}")
    print(f"Mock data only (Not Run):  {', '.join(mock_only_warehouses) or 'none'}")
    if gh_jobs:
        print(f"GH Actions data:           {', '.join(sorted(gh_jobs.keys()))}")
    if error_warehouses:
        print(f"Errors/warnings:           {', '.join(error_warehouses)}")
    else:
        print("Errors/warnings:           none")
    print("=" * 60)


if __name__ == "__main__":
    main()
