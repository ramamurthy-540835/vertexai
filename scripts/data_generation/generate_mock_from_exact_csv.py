#!/usr/bin/env python3
"""Generate leads_corrected.xlsx and pos_corrected.xlsx from the exact matching CSV.

Extracts unique leads and POS transactions from the production exact matching
output so that mock data IDs align with Cloud SQL / pipeline IDs.

Usage:
  python scripts/data_generation/generate_mock_from_exact_csv.py \
    --exact-csv reports/exact_matching/exact_matching.csv \
    --warehouse-number 115 \
    --output-dir mock_data/115

Output (local only):
  <output_dir>/leads_corrected.xlsx
  <output_dir>/pos_corrected.xlsx
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


LEAD_COLUMNS = [
    "lead_id", "warehouse_number", "fiscal_year_lead", "fiscal_period_lead",
    "week", "updated_date", "lead_source", "account_number",
    "membership_number", "lead_status", "confidence_level", "match_result",
    "type", "business_name", "address_line_one", "address_line_two",
    "city", "state", "zip_code", "email", "phone", "bd_industry",
    "industry_description", "customer_name", "closed_fiscal_period",
    "closed_fiscal_year", "batch_id", "load_date", "updated_by",
]

POS_CORE_COLUMNS = [
    "pos_id", "warehouse_number", "fiscal_year_transaction",
    "fiscal_period_transaction", "week", "account_number",
    "membership_number", "sales_reference_id", "business_name",
    "first_name", "last_name", "address_line_one", "address_line_two",
    "city", "state", "zip_code", "email", "phone", "order_amount",
    "transaction_count", "shop_type", "bd_industry", "industry_description",
    "load_date", "updated_date", "expected_relation", "expected_lead_id",
]

OMS_COLUMNS = [
    "oms_company", "oms_company_2", "oms_email_1", "oms_email_2",
    "oms_email_3", "oms_phone_1", "oms_phone_2", "oms_phone_3",
    "oms_cell_1", "oms_cell_2", "oms_first_name", "oms_middle_name",
    "oms_last_name", "oms_address_line_1", "oms_city", "oms_state",
    "oms_zip", "oms_address_line_1_v2", "oms_address_line_2",
    "oms_address_line_3", "oms_address_line_4", "oms_address_line_5",
    "oms_address_line_6", "oms_city_2", "oms_state_2", "oms_zip_2",
    "oms_zip_3", "oms_zip_4",
]

POS_COLUMNS = POS_CORE_COLUMNS + OMS_COLUMNS

LEAD_TEXT_COLS = [
    "lead_id", "warehouse_number", "lead_source", "account_number",
    "membership_number", "lead_status", "confidence_level", "match_result",
    "type", "business_name", "address_line_one", "address_line_two",
    "city", "state", "zip_code", "email", "phone", "bd_industry",
    "industry_description", "customer_name", "batch_id", "updated_by",
]

POS_TEXT_COLS = [
    "pos_id", "warehouse_number", "account_number", "membership_number",
    "sales_reference_id", "business_name", "first_name", "last_name",
    "address_line_one", "address_line_two", "city", "state", "zip_code",
    "email", "phone", "shop_type", "bd_industry", "industry_description",
    "expected_relation", "expected_lead_id",
] + OMS_COLUMNS


def format_excel_columns(path: Path, text_cols: list[str], dt_cols: list[str]) -> None:
    wb = load_workbook(path)
    ws = wb.active
    hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    for col_name in text_cols:
        if col_name not in hdr:
            continue
        ci = hdr[col_name]
        for ri in range(2, ws.max_row + 1):
            cell = ws.cell(row=ri, column=ci)
            if cell.value is not None:
                cell.number_format = "@"
    for col_name in dt_cols:
        if col_name not in hdr:
            continue
        ci = hdr[col_name]
        for ri in range(2, ws.max_row + 1):
            cell = ws.cell(row=ri, column=ci)
            if cell.value is not None:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
    wb.save(path)


def extract_leads(df: pd.DataFrame, warehouse_number: str) -> pd.DataFrame:
    """Extract unique leads from exact matching rows."""
    match_rows = df[df["match_result"].isin(["Match", "Potential"])].copy()
    ce_rows = df[df["closed_existing_flag"] == "True"].copy()

    lead_records: dict[str, dict] = {}

    for _, row in match_rows.iterrows():
        lid = row.get("lead_id", "")
        if not lid or lid in lead_records:
            continue
        lead_records[lid] = {
            "lead_id": lid,
            "warehouse_number": warehouse_number,
            "fiscal_year_lead": row.get("fiscal_year_transaction", ""),
            "fiscal_period_lead": row.get("fiscal_period_transaction", ""),
            "week": row.get("week", ""),
            "updated_date": row.get("updated_date", ""),
            "lead_source": "",
            "account_number": row.get("account_number", ""),
            "membership_number": row.get("membership_number", ""),
            "lead_status": "Open",
            "confidence_level": "",
            "match_result": row.get("match_result", ""),
            "type": "",
            "business_name": row.get("business_name_transaction", ""),
            "address_line_one": row.get("address_line_one", ""),
            "address_line_two": row.get("address_line_two", ""),
            "city": row.get("city", ""),
            "state": row.get("state", ""),
            "zip_code": row.get("zip_code", ""),
            "email": row.get("email", ""),
            "phone": row.get("phone", ""),
            "bd_industry": row.get("bd_industry", ""),
            "industry_description": row.get("industry_description", ""),
            "customer_name": "",
            "closed_fiscal_period": "",
            "closed_fiscal_year": "",
            "batch_id": "",
            "load_date": row.get("updated_date", ""),
            "updated_by": "exact_extract",
        }

    for _, row in ce_rows.iterrows():
        lid = row.get("lead_id", "")
        if not lid or lid in lead_records:
            continue
        lead_records[lid] = {
            "lead_id": lid,
            "warehouse_number": warehouse_number,
            "fiscal_year_lead": "",
            "fiscal_period_lead": "",
            "week": "",
            "updated_date": row.get("updated_date", ""),
            "lead_source": "",
            "account_number": "",
            "membership_number": "",
            "lead_status": "Closed",
            "confidence_level": "",
            "match_result": "Closed - Existing",
            "type": "",
            "business_name": "",
            "address_line_one": "",
            "address_line_two": "",
            "city": "",
            "state": "",
            "zip_code": "",
            "email": "",
            "phone": "",
            "bd_industry": "",
            "industry_description": "",
            "customer_name": "",
            "closed_fiscal_period": "",
            "closed_fiscal_year": "",
            "batch_id": "",
            "load_date": row.get("updated_date", ""),
            "updated_by": "exact_extract",
        }

    leads_df = pd.DataFrame(list(lead_records.values()), columns=LEAD_COLUMNS)
    return leads_df


def classify_relation(row: dict) -> str:
    if row.get("closed_existing_flag", "").lower() == "true":
        return "closed_existing"
    score_str = row.get("similarity_score", "")
    try:
        score = float(score_str) if score_str else 0
    except ValueError:
        score = 0
    if score >= 100:
        return "exact_single"
    if score >= 70:
        return "partial_single"
    return "unmatched"


def extract_pos(df: pd.DataFrame, warehouse_number: str) -> pd.DataFrame:
    """Extract unique POS transactions from exact matching rows."""
    match_rows = df[df["pos_id"].fillna("").str.strip() != ""].copy()

    pos_records: dict[str, dict] = {}

    for _, row in match_rows.iterrows():
        pid = row.get("pos_id", "")
        if not pid or pid in pos_records:
            continue
        biz = row.get("business_name_transaction", "")
        addr = row.get("address_line_one", "")
        city_val = row.get("city", "")
        state_val = row.get("state", "")
        zip_val = row.get("zip_code", "")
        email_val = row.get("email", "")
        phone_val = row.get("phone", "")
        first_val = row.get("first_name", "")
        last_val = row.get("last_name", "")

        rec: dict = {
            "pos_id": pid,
            "warehouse_number": warehouse_number,
            "fiscal_year_transaction": row.get("fiscal_year_transaction", ""),
            "fiscal_period_transaction": row.get("fiscal_period_transaction", ""),
            "week": row.get("week", ""),
            "account_number": row.get("account_number", ""),
            "membership_number": row.get("membership_number", ""),
            "sales_reference_id": row.get("sales_reference_id", ""),
            "business_name": biz,
            "first_name": first_val,
            "last_name": last_val,
            "address_line_one": addr,
            "address_line_two": row.get("address_line_two", ""),
            "city": city_val,
            "state": state_val,
            "zip_code": zip_val,
            "email": email_val,
            "phone": phone_val,
            "order_amount": row.get("order_amount", ""),
            "transaction_count": row.get("transaction_count", "1"),
            "shop_type": row.get("shop_type", ""),
            "bd_industry": row.get("bd_industry", ""),
            "industry_description": row.get("industry_description", ""),
            "load_date": row.get("updated_date", ""),
            "updated_date": row.get("updated_date", ""),
            "expected_relation": classify_relation(row),
            "expected_lead_id": row.get("lead_id", ""),
        }
        for col in OMS_COLUMNS:
            rec[col] = ""
        rec["oms_company"] = biz
        rec["oms_first_name"] = first_val
        rec["oms_last_name"] = last_val
        rec["oms_email_1"] = email_val
        rec["oms_phone_1"] = phone_val
        rec["oms_address_line_1"] = addr
        rec["oms_city"] = city_val
        rec["oms_state"] = state_val
        rec["oms_zip"] = zip_val

        pos_records[pid] = rec

    pos_df = pd.DataFrame(list(pos_records.values()), columns=POS_COLUMNS)
    return pos_df


def main():
    parser = argparse.ArgumentParser(
        description="Generate leads/POS xlsx from exact matching CSV"
    )
    parser.add_argument("--exact-csv", required=True,
                        help="Path to exact_matching.csv")
    parser.add_argument("--warehouse-number", required=True,
                        help="Warehouse number")
    parser.add_argument("--output-dir", required=True,
                        help="Output directory for xlsx files")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print summary without writing files")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    wh = args.warehouse_number

    print("=" * 60)
    print("Generate Mock Data from Exact Matching CSV")
    print("=" * 60)

    print(f"\nLoading: {args.exact_csv}")
    df = pd.read_csv(args.exact_csv, dtype=str, keep_default_na=False)
    print(f"  Rows: {len(df):,}")
    print(f"  Columns: {len(df.columns)}")

    print("\nExtracting leads...")
    leads_df = extract_leads(df, wh)
    match_leads = len(leads_df[leads_df["lead_status"] == "Open"])
    ce_leads = len(leads_df[leads_df["lead_status"] == "Closed"])
    print(f"  Unique leads: {len(leads_df):,} (match/potential: {match_leads:,}, CE: {ce_leads:,})")
    print(f"  Sample IDs: {leads_df['lead_id'].head(3).tolist()}")

    print("\nExtracting POS transactions...")
    pos_df = extract_pos(df, wh)
    print(f"  Unique POS: {len(pos_df):,}")
    print(f"  Sample IDs: {pos_df['pos_id'].head(3).tolist()}")
    if "expected_relation" in pos_df.columns:
        print(f"  Relation breakdown:")
        for rel, cnt in pos_df["expected_relation"].value_counts().items():
            print(f"    {rel}: {cnt:,}")

    exact_lead_ids = set(df["lead_id"].unique()) - {""}
    exact_pos_ids = set(df[df["pos_id"] != ""]["pos_id"].unique())
    mock_lead_ids = set(leads_df["lead_id"].unique())
    mock_pos_ids = set(pos_df["pos_id"].unique())

    lead_overlap = exact_lead_ids & mock_lead_ids
    pos_overlap = exact_pos_ids & mock_pos_ids

    print(f"\nAlignment check:")
    print(f"  Lead overlap: {len(lead_overlap):,}/{len(exact_lead_ids):,} ({100*len(lead_overlap)/max(len(exact_lead_ids),1):.1f}%)")
    print(f"  POS overlap:  {len(pos_overlap):,}/{len(exact_pos_ids):,} ({100*len(pos_overlap)/max(len(exact_pos_ids),1):.1f}%)")

    if args.dry_run:
        print("\n[DRY RUN] No files written.")
        return 0

    output_dir.mkdir(parents=True, exist_ok=True)
    leads_path = output_dir / "leads_corrected.xlsx"
    pos_path = output_dir / "pos_corrected.xlsx"

    print(f"\nWriting: {leads_path}")
    leads_df.to_excel(leads_path, index=False, engine="openpyxl")
    format_excel_columns(leads_path, LEAD_TEXT_COLS, ["updated_date", "load_date"])
    print(f"  {len(leads_df):,} rows")

    print(f"Writing: {pos_path}")
    pos_df.to_excel(pos_path, index=False, engine="openpyxl")
    format_excel_columns(pos_path, POS_TEXT_COLS, ["load_date", "updated_date"])
    print(f"  {len(pos_df):,} rows")

    print(f"\n{'='*60}")
    print(f"Summary")
    print(f"{'='*60}")
    print(f"  Warehouse: {wh}")
    print(f"  Source: {args.exact_csv}")
    print(f"  Leads: {len(leads_df):,} ({leads_path})")
    print(f"  POS:   {len(pos_df):,} ({pos_path})")
    print(f"  Lead alignment: {100*len(lead_overlap)/max(len(exact_lead_ids),1):.1f}%")
    print(f"  POS alignment:  {100*len(pos_overlap)/max(len(exact_pos_ids),1):.1f}%")
    print(f"\nNo Cloud SQL. No GCS. No workflow triggered.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
