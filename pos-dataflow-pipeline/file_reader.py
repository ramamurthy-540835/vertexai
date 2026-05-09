# file_reader.py
import csv, io, json, logging
from typing import List, Dict, Any

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {'.csv', '.tsv', '.txt', '.xlsx', '.xls',
                         '.parquet', '.json', '.jsonl', '.psv'}

def read_file_to_dicts(content_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    """
    Detect file type from extension and return list of row dicts.
    Supports: CSV, TSV, PSV, Excel (xls/xlsx), Parquet, JSON/JSONL.
    """
    ext = _get_ext(filename)
    logger.info(f'Reading file {filename} with extension {ext}')

    if ext in ('.csv', '.txt'):
        return _read_delimited(content_bytes, delimiter=',')
    elif ext == '.tsv':
        return _read_delimited(content_bytes, delimiter='\t')
    elif ext == '.psv':
        return _read_delimited(content_bytes, delimiter='|')
    elif ext in ('.xlsx', '.xls'):
        return _read_excel(content_bytes)
    elif ext == '.parquet':
        return _read_parquet(content_bytes)
    elif ext in ('.json',):
        return _read_json(content_bytes)
    elif ext == '.jsonl':
        return _read_jsonl(content_bytes)
    else:
        logger.warning(f'Unknown extension {ext}, attempting CSV parse')
        return _read_delimited(content_bytes, delimiter=',')

def _get_ext(filename):
    import os
    return os.path.splitext(filename.lower())[1]

def _read_delimited(content_bytes, delimiter=','):
    # Attempt UTF-8, fall back to latin-1
    for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
        try:
            text = content_bytes.decode(enc)
            # Auto-detect delimiter if not obvious
            reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
            rows = [dict(r) for r in reader]
            logger.info(f'Read {len(rows)} rows with encoding {enc}')
            return rows
        except Exception:
            continue
    raise ValueError('Could not decode file with any supported encoding')

def _read_excel(content_bytes):
    import openpyxl, xlrd
    # Try xlsx first
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else f'col_{i}'
                   for i, c in enumerate(next(ws.rows))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, [str(v) if v is not None else '' for v in row])))
        return rows
    except Exception:
        pass
    # Fall back to xlrd for .xls
    import xlrd
    wb = xlrd.open_workbook(file_contents=content_bytes)
    ws = wb.sheet_by_index(0)
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    return [dict(zip(headers, [str(ws.cell_value(r, c)) for c in range(ws.ncols)]))
            for r in range(1, ws.nrows)]

def _read_parquet(content_bytes):
    import pyarrow.parquet as pq, pyarrow as pa
    table = pq.read_table(io.BytesIO(content_bytes))
    return table.to_pydict()
    # Convert column-dict to row-dict list
    cols = table.to_pydict()
    n = len(next(iter(cols.values())))
    return [{k: str(cols[k][i]) if cols[k][i] is not None else '' for k in cols} for i in range(n)]

def _read_json(content_bytes):
    data = json.loads(content_bytes.decode('utf-8'))
    if isinstance(data, list):
        return [dict(r) for r in data]
    elif isinstance(data, dict):
        # Assume {records: [...]} or similar wrapper
        for v in data.values():
            if isinstance(v, list):
                return [dict(r) for r in v]
    raise ValueError('Cannot parse JSON as list of records')

def _read_jsonl(content_bytes):
    lines = content_bytes.decode('utf-8').splitlines()
    return [json.loads(l) for l in lines if l.strip()]
