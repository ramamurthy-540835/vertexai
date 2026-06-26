#!/usr/bin/env python3
"""Enrich existing mock-data xlsx with OMS columns and produce CSV + validation files.

Reads the existing leads_corrected.xlsx and pos_corrected.xlsx from a warehouse
mock_data directory (e.g. mock_data/115/), adds the missing OMS fields and
transaction_count column needed by the loader's transaction table, generates
additional fuzzy/CE/OAF/unmatched POS rows, and writes enriched outputs.

Output (local only, under --output-dir):
    leads_corrected.xlsx  /  leads_corrected.csv
    pos_corrected.xlsx    /  pos_corrected.csv
    expected_results.csv
    generation_summary.json
    README_generated_mock_data.md

Does NOT read from or write to Cloud SQL, GCS, SPT, or PRD.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import random
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from faker import Faker
from openpyxl import load_workbook

PERIODS_PER_YEAR = 13
CE_PERIOD_WINDOW = 6

STREET_SUFFIX_EXPAND = {
    "ST": "STREET", "AVE": "AVENUE", "RD": "ROAD",
    "DR": "DRIVE", "LN": "LANE", "BLVD": "BOULEVARD",
}
STREET_SUFFIX_ABBREV = {v: k for k, v in STREET_SUFFIX_EXPAND.items()}
STREET_SUFFIX_VARIANTS = {**STREET_SUFFIX_EXPAND, **STREET_SUFFIX_ABBREV}

INDUSTRIES = [
    ("Retail", "Retail merchandise and local storefront operations"),
    ("Food Service", "Restaurants, cafes, catering, and food preparation"),
    ("Health", "Clinics, pharmacies, and wellness providers"),
    ("Construction", "Contractors, trades, and project services"),
    ("Professional Services", "Office-based business services and consulting"),
    ("Automotive", "Vehicle sales, repair, and parts operations"),
    ("Education", "Training, tutoring, and educational services"),
    ("Real Estate", "Property management and brokerage services"),
    ("Logistics", "Delivery, storage, and logistics providers"),
    ("Manufacturing", "Light manufacturing and fabrication"),
]

OMS_COLUMNS = [
    "oms_company", "oms_company_2", "oms_email_1", "oms_email_2",
    "oms_email_3", "oms_phone_1", "oms_phone_2", "oms_phone_3",
    "oms_cell_1", "oms_cell_2", "oms_first_name", "oms_middle_name",
    "oms_last_name", "oms_address_line_1", "oms_city", "oms_state",
    "oms_zip", "oms_address_line_1_v2", "oms_address_line_2",
    "oms_address_line_3", "oms_address_line_4", "oms_address_line_5",
    "oms_address_line_6", "oms_city_2", "oms_state_2", "oms_zip_2",
    "oms_zip_3", "oms_zip_4",
]

LEAD_TEXT_COLS = [
    "lead_id", "warehouse_number", "lead_source", "account_number",
    "membership_number", "lead_status", "confidence_level", "match_result",
    "type", "business_name", "address_line_one", "address_line_two",
    "city", "state", "zip_code", "email", "phone", "bd_industry",
    "industry_description", "customer_name", "batch_id", "updated_by",
]
LEAD_DT_COLS = ["updated_date", "load_date"]

POS_TEXT_COLS = [
    "pos_id", "warehouse_number", "account_number", "membership_number",
    "sales_reference_id", "business_name", "first_name", "last_name",
    "address_line_one", "address_line_two", "city", "state", "zip_code",
    "email", "phone", "shop_type", "bd_industry", "industry_description",
    "expected_relation", "expected_lead_id",
] + OMS_COLUMNS
POS_DT_COLS = ["load_date", "updated_date"]

EXPECTED_RESULTS_COLUMNS = [
    "lead_id", "pos_id", "warehouse_number", "expected_relation",
    "expected_match_result", "expected_match_type",
    "expected_primary_transaction", "expected_closed_existing_flag",
    "expected_should_enter_fuzzy", "expected_should_be_rejected_by_fuzzy",
    "reason",
]


# ═══════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Enrich existing mock xlsx with OMS columns and produce CSVs.",
    )
    p.add_argument("--warehouse-number", type=int, required=True,
                   help="Warehouse number (resolves mock_data/<wh>/ as input).")
    p.add_argument("--input-dir", default=None,
                   help="Input directory with leads/pos xlsx (default: mock_data/<wh>/).")
    p.add_argument("--output-dir", default=None,
                   help="Output directory (default: mock_data/<wh>_from_exact/).")
    p.add_argument("--fuzzy-pos-per-lead", type=int, default=2,
                   help="Fuzzy candidate POS rows to generate per matched lead.")
    p.add_argument("--extra-unmatched", type=int, default=0,
                   help="Additional unmatched noise POS rows beyond existing.")
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--dry-run", action="store_true",
                   help="Build dataframes, validate, print summary. No file writes.")
    p.add_argument("--primary-match-csv", action="store_true",
                   help="Join leads+POS xlsx and produce primary_match_output_<uuid>.csv "
                        "in the exact pipeline output format.")
    return p.parse_args()


# ═══════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════

def safe_str(val, default: str = "") -> str:
    if val is None:
        return default
    try:
        if pd.isna(val):
            return default
    except (TypeError, ValueError):
        pass
    s = str(val).strip()
    if s.lower() in ("nan", "none", "<na>"):
        return default
    return s


def clean_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if out[col].dtype == object:
            out[col] = out[col].map(lambda v: safe_str(v))
    return out


def deterministic_id(prefix: str, *parts) -> str:
    raw = "|".join(str(p) for p in parts)
    h = hashlib.md5(raw.encode(), usedforsecurity=False).hexdigest()[:12].upper()
    return f"{prefix}{h}"


def phone_digits(rng: random.Random) -> str:
    return f"{rng.randint(200,989):03d}-{rng.randint(200,989):03d}-{rng.randint(1000,9999):04d}"


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ═══════════════════════════════════════════════════════════════
# OMS enrichment — populate OMS columns from existing POS fields
# ═══════════════════════════════════════════════════════════════

def enrich_pos_with_oms(pos_df: pd.DataFrame) -> pd.DataFrame:
    df = pos_df.copy()
    for col in OMS_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    if "transaction_count" not in df.columns:
        df["transaction_count"] = 1

    df["oms_company"] = df["oms_company"].where(
        df["oms_company"].fillna("").str.strip() != "",
        df["business_name"],
    )
    df["oms_first_name"] = df["oms_first_name"].where(
        df["oms_first_name"].fillna("").str.strip() != "",
        df.get("first_name", ""),
    )
    df["oms_last_name"] = df["oms_last_name"].where(
        df["oms_last_name"].fillna("").str.strip() != "",
        df.get("last_name", ""),
    )
    df["oms_address_line_1"] = df["oms_address_line_1"].where(
        df["oms_address_line_1"].fillna("").str.strip() != "",
        df.get("address_line_one", ""),
    )
    df["oms_city"] = df["oms_city"].where(
        df["oms_city"].fillna("").str.strip() != "",
        df.get("city", ""),
    )
    df["oms_state"] = df["oms_state"].where(
        df["oms_state"].fillna("").str.strip() != "",
        df.get("state", ""),
    )
    df["oms_zip"] = df["oms_zip"].where(
        df["oms_zip"].fillna("").str.strip() != "",
        df.get("zip_code", ""),
    )
    df["oms_email_1"] = df["oms_email_1"].where(
        df["oms_email_1"].fillna("").str.strip() != "",
        df.get("email", ""),
    )
    df["oms_phone_1"] = df["oms_phone_1"].where(
        df["oms_phone_1"].fillna("").str.strip() != "",
        df.get("phone", ""),
    )
    return df


# ═══════════════════════════════════════════════════════════════
# Additional POS row generators (fuzzy, CE, OAF, unmatched)
# ═══════════════════════════════════════════════════════════════

def mutate_business_name(name: str, rng: random.Random) -> str:
    if not name:
        return name
    parts = name.split()
    mutations = [
        lambda p: [w.replace("PRESCHOOL", "PRE SCHOOL") if "PRESCHOOL" in w.upper() else w for w in p],
        lambda p: [w for w in p if w.upper() not in ("LLC", "INC", "INC.", "CO", "CO.")],
        lambda p: p + [rng.choice(["LLC", "Inc", "Co"])],
        lambda p: [w for w in p if w.upper() not in ("THE", "AND", "&")],
        lambda p: [w.upper() for w in p],
        lambda p: [w.title() for w in p],
    ]
    chosen = rng.choice(mutations)
    result = " ".join(w for w in chosen(list(parts)) if w.strip())
    if result == name and len(parts) > 1:
        idx = rng.randrange(len(parts))
        parts[idx] = parts[idx][:max(3, len(parts[idx]) - 2)]
        result = " ".join(parts)
    return result or name


def mutate_address(address: str, rng: random.Random) -> str:
    if not address:
        return address
    parts = address.upper().split()
    for i, part in enumerate(parts):
        if part in STREET_SUFFIX_VARIANTS:
            parts[i] = STREET_SUFFIX_VARIANTS[part]
            break
    roll = rng.random()
    if roll < 0.3:
        parts.append(f"STE {rng.randint(100, 999)}")
    elif roll < 0.5:
        parts.append(f"UNIT {rng.randint(1, 50)}")
    if rng.random() < 0.3:
        parts = [p.title() for p in parts]
    return " ".join(parts)


def _shift_period_back(fy: int, fp: int, n: int) -> tuple[int, int]:
    pos_fp = fp - n
    pos_fy = fy
    while pos_fp < 1:
        pos_fp += PERIODS_PER_YEAR
        pos_fy -= 1
    return pos_fy, pos_fp


def _base_pos_dict(wh: int, pos_id: str, sr_id: str) -> dict:
    d: dict = {
        "pos_id": pos_id,
        "warehouse_number": str(wh),
        "transaction_count": 1,
        "load_date": now_str(),
        "updated_date": now_str(),
    }
    for c in OMS_COLUMNS:
        d[c] = ""
    return d


def generate_fuzzy_row(lead: pd.Series, wh: int, rng: random.Random, seq: int) -> dict:
    lid = safe_str(lead.get("lead_id"))
    biz = mutate_business_name(safe_str(lead.get("business_name")), rng)
    addr = mutate_address(safe_str(lead.get("address_line_one")), rng)
    fy = lead.get("fiscal_year_lead", 2026)
    fp = lead.get("fiscal_period_lead", 1)
    try:
        fy = int(fy)
    except (ValueError, TypeError):
        fy = 2026
    try:
        fp = int(fp)
    except (ValueError, TypeError):
        fp = 1
    pos_fp = min(PERIODS_PER_YEAR, fp + rng.randint(0, 3))

    email = safe_str(lead.get("email"))
    if email and rng.random() < 0.5:
        email = ""
    phone = safe_str(lead.get("phone"))
    if phone and rng.random() < 0.6:
        phone = ""

    r = _base_pos_dict(wh, deterministic_id("POS-FZ-", wh, lid, seq),
                       deterministic_id("SR-FZ-", wh, lid, seq))
    r.update({
        "fiscal_year_transaction": fy,
        "fiscal_period_transaction": pos_fp,
        "week": rng.randint(1, 5),
        "account_number": safe_str(lead.get("account_number")),
        "membership_number": safe_str(lead.get("membership_number")),
        "sales_reference_id": r.pop("sales_reference_id", deterministic_id("SR-FZ-", wh, lid, seq)),
        "business_name": biz,
        "first_name": "",
        "last_name": "",
        "address_line_one": addr,
        "address_line_two": "",
        "city": safe_str(lead.get("city")),
        "state": safe_str(lead.get("state")),
        "zip_code": safe_str(lead.get("zip_code")),
        "email": email,
        "phone": phone,
        "order_amount": round(rng.uniform(50, 3000), 2),
        "shop_type": rng.choice(["Delivery", "In-Warehouse", "Pickup"]),
        "bd_industry": safe_str(lead.get("bd_industry")),
        "industry_description": safe_str(lead.get("industry_description")),
        "expected_relation": "fuzzy_candidate",
        "expected_lead_id": lid,
    })
    r["sales_reference_id"] = deterministic_id("SR-FZ-", wh, lid, seq)
    r["oms_company"] = biz
    if addr:
        r["oms_address_line_1"] = addr
        r["oms_city"] = safe_str(lead.get("city"))
        r["oms_state"] = safe_str(lead.get("state"))
        r["oms_zip"] = safe_str(lead.get("zip_code"))
    return r


def generate_ce_row(lead: pd.Series, wh: int, rng: random.Random, seq: int) -> dict:
    lid = safe_str(lead.get("lead_id"))
    fy = int(lead.get("fiscal_year_lead", 2026) or 2026)
    fp = int(lead.get("fiscal_period_lead", 1) or 1)
    pos_fy, pos_fp = _shift_period_back(fy, fp, rng.randint(1, CE_PERIOD_WINDOW))

    r = _base_pos_dict(wh, deterministic_id("POS-CE-", wh, lid, seq),
                       deterministic_id("SR-CE-", wh, lid, seq))
    r.update({
        "fiscal_year_transaction": pos_fy,
        "fiscal_period_transaction": pos_fp,
        "week": rng.randint(1, 5),
        "account_number": safe_str(lead.get("account_number")),
        "membership_number": safe_str(lead.get("membership_number")),
        "sales_reference_id": deterministic_id("SR-CE-", wh, lid, seq),
        "business_name": safe_str(lead.get("business_name")),
        "first_name": "",
        "last_name": "",
        "address_line_one": safe_str(lead.get("address_line_one")),
        "address_line_two": "",
        "city": safe_str(lead.get("city")),
        "state": safe_str(lead.get("state")),
        "zip_code": safe_str(lead.get("zip_code")),
        "email": safe_str(lead.get("email")),
        "phone": safe_str(lead.get("phone")),
        "order_amount": round(rng.uniform(30, 1500), 2),
        "shop_type": rng.choice(["Delivery", "In-Warehouse", "Pickup"]),
        "bd_industry": safe_str(lead.get("bd_industry")),
        "industry_description": safe_str(lead.get("industry_description")),
        "expected_relation": "closed_existing",
        "expected_lead_id": lid,
    })
    r["oms_company"] = safe_str(lead.get("business_name"))
    return r


def generate_oaf_row(lead: pd.Series, wh: int, rng: random.Random, seq: int) -> dict:
    lid = safe_str(lead.get("lead_id"))
    fy = int(lead.get("fiscal_year_lead", 2026) or 2026)
    fp = int(lead.get("fiscal_period_lead", 1) or 1)
    pos_fy, pos_fp = _shift_period_back(fy, fp, rng.randint(CE_PERIOD_WINDOW + 1, CE_PERIOD_WINDOW + 6))

    r = _base_pos_dict(wh, deterministic_id("POS-OAF-", wh, lid, seq),
                       deterministic_id("SR-OAF-", wh, lid, seq))
    r.update({
        "fiscal_year_transaction": pos_fy,
        "fiscal_period_transaction": pos_fp,
        "week": rng.randint(1, 5),
        "account_number": safe_str(lead.get("account_number")),
        "membership_number": safe_str(lead.get("membership_number")),
        "sales_reference_id": deterministic_id("SR-OAF-", wh, lid, seq),
        "business_name": safe_str(lead.get("business_name")),
        "first_name": "",
        "last_name": "",
        "address_line_one": safe_str(lead.get("address_line_one")),
        "address_line_two": "",
        "city": safe_str(lead.get("city")),
        "state": safe_str(lead.get("state")),
        "zip_code": safe_str(lead.get("zip_code")),
        "email": safe_str(lead.get("email")),
        "phone": safe_str(lead.get("phone")),
        "order_amount": round(rng.uniform(30, 1500), 2),
        "shop_type": rng.choice(["Delivery", "In-Warehouse", "Pickup"]),
        "bd_industry": safe_str(lead.get("bd_industry")),
        "industry_description": safe_str(lead.get("industry_description")),
        "expected_relation": "oaf_drop",
        "expected_lead_id": lid,
    })
    return r


def generate_unmatched_row(wh: int, rng: random.Random, fake: Faker, seq: int) -> dict:
    biz = fake.company().replace(",", "")
    addr = fake.street_address().replace(",", "")
    city = fake.city()
    state = fake.state_abbr().upper()
    zc = fake.zipcode()[:5].zfill(5)
    ind, desc = INDUSTRIES[rng.randrange(len(INDUSTRIES))]

    email_local = re.sub(r"[^a-z0-9]+", ".", biz.lower()).strip(".")
    email = f"{email_local}@{fake.domain_name()}".replace("..", ".")

    r = _base_pos_dict(wh, deterministic_id("POS-UM-", wh, seq, rng.randint(0, 99999)),
                       deterministic_id("SR-UM-", wh, seq, rng.randint(0, 99999)))
    r.update({
        "fiscal_year_transaction": rng.choice([2025, 2026]),
        "fiscal_period_transaction": rng.randint(1, PERIODS_PER_YEAR),
        "week": rng.randint(1, 5),
        "account_number": str(rng.randint(10000000, 99999999)),
        "membership_number": str(rng.randint(1000000000, 9999999999)),
        "sales_reference_id": deterministic_id("SR-UM-", wh, seq, rng.randint(0, 99999)),
        "business_name": biz,
        "first_name": fake.first_name(),
        "last_name": fake.last_name(),
        "address_line_one": addr,
        "address_line_two": rng.choice(["", "Unit 2", "Suite 100", "Bldg A"]),
        "city": city,
        "state": state,
        "zip_code": zc,
        "email": email if rng.random() < 0.7 else "",
        "phone": phone_digits(rng) if rng.random() < 0.8 else "",
        "order_amount": round(rng.uniform(19, 2500), 2),
        "shop_type": rng.choice(["Delivery", "In-Warehouse", "Pickup"]),
        "bd_industry": ind,
        "industry_description": desc,
        "expected_relation": "unmatched",
        "expected_lead_id": "",
    })
    r["oms_company"] = biz
    r["oms_first_name"] = r["first_name"]
    r["oms_last_name"] = r["last_name"]
    return r


# ═══════════════════════════════════════════════════════════════
# Expected results
# ═══════════════════════════════════════════════════════════════

_RELATION_EXPECTED = {
    "exact_single": {
        "mr": "Match", "mt": "Exact", "ce": False,
        "fuzzy": False, "rej": False,
        "reason": "Exact single-match POS from existing mock generator.",
    },
    "exact_multi": {
        "mr": "Match", "mt": "Exact", "ce": False,
        "fuzzy": False, "rej": False,
        "reason": "Exact multi-match POS from existing mock generator.",
    },
    "partial_single": {
        "mr": "Potential", "mt": "Fuzzy", "ce": False,
        "fuzzy": True, "rej": False,
        "reason": "Partial single-match POS from existing mock generator.",
    },
    "partial_multi": {
        "mr": "Potential", "mt": "Fuzzy", "ce": False,
        "fuzzy": True, "rej": False,
        "reason": "Partial multi-match POS from existing mock generator.",
    },
    "closed_existing_only": {
        "mr": "Closed - Existing", "mt": "", "ce": True,
        "fuzzy": False, "rej": False,
        "reason": "Closed-existing POS from existing mock generator.",
    },
    "closed_existing": {
        "mr": "Closed - Existing", "mt": "", "ce": True,
        "fuzzy": False, "rej": False,
        "reason": "Pre-lead POS within 6-period CE window.",
    },
    "fuzzy_candidate": {
        "mr": "Potential", "mt": "Fuzzy", "ce": False,
        "fuzzy": True, "rej": False,
        "reason": "Name/address variation of lead; should score 70-99.999 in fuzzy.",
    },
    "oaf_drop": {
        "mr": "", "mt": "", "ce": False,
        "fuzzy": False, "rej": False,
        "reason": "Pre-lead POS older than 6 periods; silently dropped (OAF).",
    },
    "unmatched": {
        "mr": "", "mt": "", "ce": False,
        "fuzzy": True, "rej": True,
        "reason": "Unrelated noise POS; should not match any lead.",
    },
    "no_match": {
        "mr": "", "mt": "", "ce": False,
        "fuzzy": True, "rej": True,
        "reason": "Noise POS from existing mock generator; no match expected.",
    },
    "late_cycle_unmatched": {
        "mr": "", "mt": "", "ce": False,
        "fuzzy": True, "rej": True,
        "reason": "Late-cycle unmatched POS from existing mock generator.",
    },
}


def build_expected_row(pos_row: dict) -> dict:
    rel = safe_str(pos_row.get("expected_relation"), "unmatched")
    info = _RELATION_EXPECTED.get(rel, _RELATION_EXPECTED["unmatched"])
    return {
        "lead_id": safe_str(pos_row.get("expected_lead_id")),
        "pos_id": safe_str(pos_row.get("pos_id")),
        "warehouse_number": safe_str(pos_row.get("warehouse_number")),
        "expected_relation": rel,
        "expected_match_result": info["mr"],
        "expected_match_type": info["mt"],
        "expected_primary_transaction": "",
        "expected_closed_existing_flag": info["ce"],
        "expected_should_enter_fuzzy": info["fuzzy"],
        "expected_should_be_rejected_by_fuzzy": info["rej"],
        "reason": info["reason"],
    }


def mark_primary_transactions(pos_df: pd.DataFrame, expected_rows: list[dict]) -> None:
    exact_rels = {"exact_single", "exact_multi"}
    exact_mask = pos_df["expected_relation"].isin(exact_rels)
    if not exact_mask.any():
        return

    exact = pos_df[exact_mask].copy()
    primary_ids: set[str] = set()
    for lid, grp in exact.groupby("expected_lead_id"):
        fy = pd.to_numeric(grp["fiscal_year_transaction"], errors="coerce").fillna(9999)
        fp = pd.to_numeric(grp["fiscal_period_transaction"], errors="coerce").fillna(99)
        wk = pd.to_numeric(grp["week"], errors="coerce").fillna(99)
        order = fy * 10000 + fp * 100 + wk
        primary_ids.add(str(grp.loc[order.idxmin(), "pos_id"]))

    for er in expected_rows:
        if er["pos_id"] in primary_ids:
            er["expected_primary_transaction"] = True


# ═══════════════════════════════════════════════════════════════
# Deduplication
# ═══════════════════════════════════════════════════════════════

def deduplicate_ids(rows: list[dict]) -> list[dict]:
    seen_pos: set[str] = set()
    seen_sr: set[str] = set()
    out: list[dict] = []
    for row in rows:
        pid = str(row.get("pos_id", ""))
        sid = str(row.get("sales_reference_id", ""))
        sfx = 0
        orig_pid, orig_sid = pid, sid
        while pid in seen_pos:
            sfx += 1
            pid = f"{orig_pid}-{sfx}"
        while sid in seen_sr:
            sfx += 1
            sid = f"{orig_sid}-{sfx}"
        row["pos_id"] = pid
        row["sales_reference_id"] = sid
        seen_pos.add(pid)
        seen_sr.add(sid)
        out.append(row)
    return out


# ═══════════════════════════════════════════════════════════════
# Validation
# ═══════════════════════════════════════════════════════════════

def validate_no_duplicate_pos_id(df: pd.DataFrame) -> int:
    dupes = int(df["pos_id"].duplicated().sum())
    print(f"  {'WARN' if dupes else 'OK  '} duplicate pos_id: {dupes}")
    return dupes


def validate_no_duplicate_sales_reference_id(df: pd.DataFrame) -> int:
    dupes = int(df["sales_reference_id"].duplicated().sum())
    print(f"  {'WARN' if dupes else 'OK  '} duplicate sales_reference_id: {dupes}")
    return dupes


def validate_primary_transaction(edf: pd.DataFrame) -> int:
    exact = edf[edf["expected_relation"].isin(["exact_single", "exact_multi"])]
    if exact.empty:
        print("  OK   no exact-match rows (primary_transaction N/A)")
        return 0
    leads = set(exact["lead_id"].unique())
    primary = set(exact[exact["expected_primary_transaction"] == True]["lead_id"].unique())  # noqa: E712
    missing = len(leads - primary)
    print(f"  {'WARN' if missing else 'OK  '} exact leads missing primary_transaction: {missing}")
    return missing


def validate_warehouse_scope(df: pd.DataFrame, wh: int) -> int:
    wrong = len(df[df["warehouse_number"].astype(str) != str(wh)])
    print(f"  {'WARN' if wrong else 'OK  '} POS rows outside warehouse {wh}: {wrong}")
    return wrong


def validate_expected_results_consistency(edf: pd.DataFrame) -> int:
    issues = 0
    for rel in ["exact_single", "exact_multi"]:
        sub = edf[edf["expected_relation"] == rel]
        issues += int((sub["expected_match_type"] != "Exact").sum())
    ce = edf[edf["expected_relation"].isin(["closed_existing", "closed_existing_only"])]
    issues += int((ce["expected_closed_existing_flag"] != True).sum())  # noqa: E712
    um = edf[edf["expected_relation"].isin(["unmatched", "no_match", "late_cycle_unmatched"])]
    issues += int((um["expected_should_be_rejected_by_fuzzy"] != True).sum())  # noqa: E712
    print(f"  {'WARN' if issues else 'OK  '} expected_results consistency issues: {issues}")
    return issues


def validate_no_nan_strings(df: pd.DataFrame, label: str) -> int:
    count = 0
    for col in df.columns:
        if df[col].dtype == object:
            count += int(df[col].fillna("").astype(str).str.lower().isin(["nan", "none", "<na>"]).sum())
    print(f"  {'WARN' if count else 'OK  '} [{label}] string nan/none values: {count}")
    return count


def validate_required_columns(df: pd.DataFrame, required: list[str], label: str) -> list[str]:
    missing = [c for c in required if c not in df.columns]
    print(f"  {'WARN' if missing else 'OK  '} [{label}] columns present: {len(required) - len(missing)}/{len(required)}"
          + (f" missing: {missing}" if missing else ""))
    return missing


# ═══════════════════════════════════════════════════════════════
# Excel / CSV writers
# ═══════════════════════════════════════════════════════════════

def format_excel_columns(path: Path, text_cols: list[str], dt_cols: list[str]) -> None:
    wb = load_workbook(path)
    ws = wb.active
    hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    for col_name in text_cols:
        if col_name not in hdr:
            continue
        ci = hdr[col_name]
        for ri in range(2, ws.max_row + 1):
            cell = ws.cell(row=ri, column=ci)
            if cell.value is not None:
                cell.number_format = "@"
    for col_name in dt_cols:
        if col_name not in hdr:
            continue
        ci = hdr[col_name]
        for ri in range(2, ws.max_row + 1):
            cell = ws.cell(row=ri, column=ci)
            if cell.value is not None:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
    wb.save(path)


def write_workbook(df: pd.DataFrame, path: Path, text_cols: list[str], dt_cols: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False, engine="openpyxl")
    format_excel_columns(path, text_cols, dt_cols)


def write_csv_safe(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    out = df.copy()
    for col in out.columns:
        if out[col].dtype == object:
            out[col] = out[col].fillna("").astype(str)
    out.to_csv(path, index=False)


# ═══════════════════════════════════════════════════════════════
# Primary match output CSV — pipeline exact-output format
# ═══════════════════════════════════════════════════════════════

PRIMARY_MATCH_HEADER = [
    "lead_id", "pos_id", "match_result", "similarity_score", "winning_set",
    "match_type", "primary_transaction", "matched_by", "matching_comments",
    "closed_existing_flag", "account_number", "transaction_count",
    "business_name_transaction", "membership_number", "warehouse_number",
    "sales_reference_id", "fiscal_year_transaction", "fiscal_period_transaction",
    "week", "shop_type", "bd_industry", "order_amount", "industry_description",
    "first_name", "last_name", "address_line_one", "address_line_two",
    "city", "state", "zip_code", "email", "phone",
    "u_matched_lead_number", "u_order_amount", "u_order_amount_rounded",
    "updated_date",
]


def build_primary_match_csv(leads_df: pd.DataFrame, pos_df: pd.DataFrame,
                            wh: int, output_dir: Path) -> Path:
    import uuid as _uuid

    leads = leads_df.copy()
    pos = pos_df.copy()

    leads = clean_df(leads)
    pos = clean_df(pos)

    matched = pos[pos["expected_lead_id"].fillna("").str.strip() != ""].copy()
    if matched.empty:
        print("ERROR: No matched POS rows (expected_lead_id is empty for all rows).")
        raise SystemExit(1)

    merged = matched.merge(
        leads[["lead_id", "business_name"]].rename(columns={"business_name": "lead_biz_name"}),
        left_on="expected_lead_id",
        right_on="lead_id",
        how="left",
        suffixes=("", "_lead"),
    )

    merged["_fy"] = pd.to_numeric(merged["fiscal_year_transaction"], errors="coerce").fillna(9999)
    merged["_fp"] = pd.to_numeric(merged["fiscal_period_transaction"], errors="coerce").fillna(99)
    merged["_wk"] = pd.to_numeric(merged["week"], errors="coerce").fillna(99)
    merged["_order"] = merged["_fy"] * 10000 + merged["_fp"] * 100 + merged["_wk"]

    primary_idx = merged.groupby("expected_lead_id")["_order"].idxmin()
    merged["primary_transaction"] = False
    merged.loc[primary_idx, "primary_transaction"] = True

    order_amt = pd.to_numeric(merged["order_amount"], errors="coerce").fillna(0.0)

    def _match_comment(row):
        parts = [f"Complete Match (score 150/150). Fields matched: business_name"]
        if safe_str(row.get("email")):
            parts.append("email")
        if safe_str(row.get("phone")):
            parts.append("phone")
        if safe_str(row.get("zip_code")):
            parts.append("zip_code")
        if safe_str(row.get("city")):
            parts.append("city")
        if safe_str(row.get("state")):
            parts.append("state")
        base = ", ".join(parts) + "."
        if row.get("primary_transaction"):
            base += " Designated as primary transaction (earliest fiscal period for this lead)."
        return base

    rows: list[dict] = []
    for _, r in merged.iterrows():
        oa = float(order_amt.loc[r.name]) if r.name in order_amt.index else 0.0
        rows.append({
            "lead_id": safe_str(r.get("expected_lead_id")),
            "pos_id": safe_str(r.get("pos_id")),
            "match_result": "Match",
            "similarity_score": 150.0,
            "winning_set": 4.0,
            "match_type": "Exact",
            "primary_transaction": r.get("primary_transaction", False),
            "matched_by": "System",
            "matching_comments": _match_comment(r),
            "closed_existing_flag": False,
            "account_number": safe_str(r.get("account_number")),
            "transaction_count": 1.0,
            "business_name_transaction": safe_str(r.get("business_name")),
            "membership_number": safe_str(r.get("membership_number")),
            "warehouse_number": str(wh),
            "sales_reference_id": safe_str(r.get("sales_reference_id")),
            "fiscal_year_transaction": safe_str(r.get("fiscal_year_transaction")),
            "fiscal_period_transaction": safe_str(r.get("fiscal_period_transaction")),
            "week": safe_str(r.get("week")),
            "shop_type": safe_str(r.get("shop_type")),
            "bd_industry": safe_str(r.get("bd_industry")),
            "order_amount": oa,
            "industry_description": safe_str(r.get("industry_description")),
            "first_name": safe_str(r.get("first_name")),
            "last_name": safe_str(r.get("last_name")),
            "address_line_one": safe_str(r.get("address_line_one")),
            "address_line_two": safe_str(r.get("address_line_two")),
            "city": safe_str(r.get("city")),
            "state": safe_str(r.get("state")),
            "zip_code": safe_str(r.get("zip_code")),
            "email": safe_str(r.get("email")),
            "phone": safe_str(r.get("phone")),
            "u_matched_lead_number": safe_str(r.get("expected_lead_id")),
            "u_order_amount": oa,
            "u_order_amount_rounded": round(oa, 2),
            "updated_date": safe_str(r.get("updated_date")) or now_str(),
        })

    out_df = pd.DataFrame(rows, columns=PRIMARY_MATCH_HEADER)

    run_id = str(_uuid.uuid4())
    filename = f"primary_match_output_{run_id}.csv"
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / filename

    for col in out_df.columns:
        if out_df[col].dtype == object:
            out_df[col] = out_df[col].fillna("").astype(str)
    out_df.to_csv(out_path, index=False)

    print(f"\nPrimary match output CSV:")
    print(f"  File: {out_path}")
    print(f"  Rows: {len(out_df):,}")
    print(f"  Leads: {out_df['lead_id'].nunique():,}")
    print(f"  Primary transactions: {out_df['primary_transaction'].sum():,}")
    print(f"  Warehouse: {wh}")
    return out_path


# ═══════════════════════════════════════════════════════════════
# main
# ═══════════════════════════════════════════════════════════════

def main() -> int:
    args = parse_args()
    wh = args.warehouse_number

    script_dir = Path(__file__).resolve().parent
    input_dir = Path(args.input_dir) if args.input_dir else script_dir / str(wh)
    output_dir = Path(args.output_dir) if args.output_dir else script_dir / f"{wh}_from_exact"

    leads_path = input_dir / "leads_corrected.xlsx"
    pos_path = input_dir / "pos_corrected.xlsx"

    if not leads_path.exists():
        print(f"ERROR: {leads_path} not found")
        return 1
    if not pos_path.exists():
        print(f"ERROR: {pos_path} not found")
        return 1

    # ── Primary match CSV mode ─────────────────────────────────
    if args.primary_match_csv:
        print(f"Reading input from: {input_dir}/")
        leads_df = pd.read_excel(leads_path)
        pos_df = pd.read_excel(pos_path)
        print(f"  Leads: {len(leads_df):,} rows")
        print(f"  POS:   {len(pos_df):,} rows")
        out_path = build_primary_match_csv(leads_df, pos_df, wh, output_dir)
        print(f"\nTo upload to GCS:")
        print(f"  gcloud storage cp {out_path} gs://lead-match-ctoteam/match/final_match_result/{out_path.name}")
        return 0

    rng = random.Random(args.seed)
    Faker.seed(args.seed)
    fake = Faker("en_US")
    fake.seed_instance(args.seed)

    # ── Read existing xlsx ─────────────────────────────────────
    print(f"Reading input from: {input_dir}/")
    leads_df = pd.read_excel(leads_path)
    pos_df = pd.read_excel(pos_path)
    print(f"  Leads: {len(leads_df):,} rows, {len(leads_df.columns)} columns")
    print(f"  POS:   {len(pos_df):,} rows, {len(pos_df.columns)} columns")

    # ── Clean nan strings ──────────────────────────────────────
    leads_df = clean_df(leads_df)
    pos_df = clean_df(pos_df)

    # ── Enrich POS with OMS columns + transaction_count ────────
    pos_df = enrich_pos_with_oms(pos_df)
    print(f"  POS after OMS enrichment: {len(pos_df.columns)} columns")

    # ── Identify matched leads for extra row generation ────────
    matched_rels = {"exact_single", "exact_multi", "partial_single", "partial_multi", "closed_existing_only"}
    if "expected_relation" in pos_df.columns and "expected_lead_id" in pos_df.columns:
        matched_pos = pos_df[pos_df["expected_relation"].isin(matched_rels)]
        matched_lead_ids = matched_pos["expected_lead_id"].dropna().unique()
        matched_leads = leads_df[leads_df["lead_id"].isin(matched_lead_ids)]
    else:
        matched_leads = leads_df

    print(f"  Matched leads for extra generation: {len(matched_leads):,}")

    # ── Generate additional POS rows ───────────────────────────
    new_rows: list[dict] = []

    ce_frac, oaf_frac = 0.15, 0.08
    n_leads = len(matched_leads)
    ce_n = max(1, int(n_leads * ce_frac))
    oaf_n = max(1, int(n_leads * oaf_frac))
    lead_indices = list(range(n_leads))
    rng.shuffle(lead_indices)
    ce_indices = set(lead_indices[:ce_n])
    oaf_indices = set(lead_indices[ce_n: ce_n + oaf_n])

    for idx, (_, lead) in enumerate(matched_leads.iterrows()):
        for seq in range(1, args.fuzzy_pos_per_lead + 1):
            new_rows.append(generate_fuzzy_row(lead, wh, rng, seq))
        if idx in ce_indices:
            new_rows.append(generate_ce_row(lead, wh, rng, 1))
        if idx in oaf_indices:
            new_rows.append(generate_oaf_row(lead, wh, rng, 1))

    for seq in range(1, args.extra_unmatched + 1):
        new_rows.append(generate_unmatched_row(wh, rng, fake, seq))

    new_rows = deduplicate_ids(new_rows)

    # ── Merge new rows into POS df ─────────────────────────────
    if new_rows:
        new_df = pd.DataFrame(new_rows)
        for col in pos_df.columns:
            if col not in new_df.columns:
                new_df[col] = ""
        for col in new_df.columns:
            if col not in pos_df.columns:
                pos_df[col] = ""
        pos_df = pd.concat([pos_df, new_df[pos_df.columns]], ignore_index=True)

    pos_df = clean_df(pos_df)
    pos_df["warehouse_number"] = pos_df["warehouse_number"].astype(str)
    pos_df["zip_code"] = pos_df["zip_code"].astype(str)
    leads_df["warehouse_number"] = leads_df["warehouse_number"].astype(str)
    leads_df["zip_code"] = leads_df["zip_code"].astype(str)

    # ── Count by relation ──────────────────────────────────────
    relation_counts = pos_df["expected_relation"].value_counts().to_dict() if "expected_relation" in pos_df.columns else {}
    new_by_rel: dict[str, int] = {}
    for r in new_rows:
        rel = r.get("expected_relation", "unknown")
        new_by_rel[rel] = new_by_rel.get(rel, 0) + 1

    counts = {
        "warehouse_number": wh,
        "input_leads": len(leads_df),
        "input_pos_original": len(pos_df) - len(new_rows),
        "new_fuzzy_candidate": new_by_rel.get("fuzzy_candidate", 0),
        "new_closed_existing": new_by_rel.get("closed_existing", 0),
        "new_oaf_drop": new_by_rel.get("oaf_drop", 0),
        "new_unmatched": new_by_rel.get("unmatched", 0),
        "total_pos_output": len(pos_df),
        "oms_columns_added": len(OMS_COLUMNS),
        "seed": args.seed,
    }
    counts["relation_breakdown"] = relation_counts

    # ── Build expected results ─────────────────────────────────
    expected_rows = [build_expected_row(row.to_dict()) for _, row in pos_df.iterrows()]
    mark_primary_transactions(pos_df, expected_rows)
    expected_df = pd.DataFrame(expected_rows, columns=EXPECTED_RESULTS_COLUMNS)

    # ── Validation ─────────────────────────────────────────────
    print(f"\nValidation:")
    validate_required_columns(leads_df, [c for c in LEAD_TEXT_COLS] + LEAD_DT_COLS, "leads")
    pos_expected_cols = POS_TEXT_COLS + POS_DT_COLS + ["order_amount", "transaction_count"]
    validate_required_columns(pos_df, list(dict.fromkeys(pos_expected_cols)), "pos")
    validate_no_duplicate_pos_id(pos_df)
    validate_no_duplicate_sales_reference_id(pos_df)
    validate_primary_transaction(expected_df)
    validate_warehouse_scope(pos_df, wh)
    validate_expected_results_consistency(expected_df)
    validate_no_nan_strings(leads_df, "leads")
    validate_no_nan_strings(pos_df, "pos")

    # ── Summary ────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"Generation Summary — warehouse {wh}")
    print(f"{'='*60}")
    print(f"  Input:  {input_dir}/")
    print(f"  Output: {output_dir}/")
    for k, v in counts.items():
        if k == "relation_breakdown":
            print(f"  {k}:")
            for rk, rv in v.items():
                print(f"    {rk}: {rv:,}")
        elif isinstance(v, int):
            print(f"  {k}: {v:,}")
        else:
            print(f"  {k}: {v}")
    print(f"{'='*60}")

    if args.dry_run:
        print("\n[DRY RUN] No files written.")
        return 0

    # ── Write output files ─────────────────────────────────────
    output_dir.mkdir(parents=True, exist_ok=True)

    leads_xlsx = output_dir / "leads_corrected.xlsx"
    leads_csv = output_dir / "leads_corrected.csv"
    pos_xlsx = output_dir / "pos_corrected.xlsx"
    pos_csv = output_dir / "pos_corrected.csv"
    expected_p = output_dir / "expected_results.csv"
    summary_p = output_dir / "generation_summary.json"
    readme_p = output_dir / "README_generated_mock_data.md"

    write_workbook(leads_df, leads_xlsx, LEAD_TEXT_COLS, LEAD_DT_COLS)
    write_csv_safe(leads_df, leads_csv)
    write_workbook(pos_df, pos_xlsx, POS_TEXT_COLS, POS_DT_COLS)
    write_csv_safe(pos_df, pos_csv)
    expected_df.to_csv(expected_p, index=False)

    summary_json = {
        **{k: v for k, v in counts.items() if k != "relation_breakdown"},
        "relation_breakdown": {str(k): int(v) for k, v in relation_counts.items()},
        "generated_at": datetime.now().isoformat(),
        "input_dir": str(input_dir),
        "output_dir": str(output_dir),
        "files": [str(p) for p in [leads_xlsx, leads_csv, pos_xlsx, pos_csv, expected_p]],
    }
    with open(summary_p, "w") as f:
        json.dump(summary_json, f, indent=2, default=str)

    readme = f"""# Generated Mock Data — Warehouse {wh}

Enriched from existing mock data using `generate_from_exact_output.py`.

## Source
- **Input:** `{input_dir}/`
- **Warehouse:** {wh}
- **Seed:** {args.seed}

## Files
| File | Rows | Description |
|------|------|-------------|
| leads_corrected.xlsx | {len(leads_df):,} | Lead workbook (Excel) |
| leads_corrected.csv | {len(leads_df):,} | Lead workbook (CSV) |
| pos_corrected.xlsx | {len(pos_df):,} | POS workbook with OMS fields (Excel) |
| pos_corrected.csv | {len(pos_df):,} | POS workbook with OMS fields (CSV) |
| expected_results.csv | {len(expected_df):,} | Per-POS validation truth |
| generation_summary.json | — | Machine-readable summary |

## POS Composition
| Category | Count |
|----------|-------|
"""
    for rel, cnt in sorted(relation_counts.items()):
        readme += f"| {rel} | {cnt:,} |\n"

    readme += f"""
## What changed
- Added {len(OMS_COLUMNS)} OMS columns (oms_company, oms_email_1, oms_address_line_1, etc.)
- Added transaction_count column
- Generated fuzzy_candidate, closed_existing, oaf_drop rows from matched leads
- Produced CSV versions of both Excel files
- Created expected_results.csv for pipeline validation

## Safety
- No Cloud SQL load was performed.
- No GCS upload was performed.
- No SPT or PRD resources were touched.
"""
    with open(readme_p, "w") as f:
        f.write(readme)

    print(f"\nOutput written to: {output_dir}/")
    for p in [leads_xlsx, leads_csv, pos_xlsx, pos_csv, expected_p, summary_p, readme_p]:
        print(f"  {p.name}")
    print(f"\nNo Cloud SQL. No GCS. No SPT/PRD.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
