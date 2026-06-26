#!/usr/bin/env python3
"""Generate realistic mock leads/POS Excel files for one Costco warehouse.

The generated workbooks are designed to exercise the lead-to-POS matching
pipeline with a controlled mix of:
  - exact matches
  - partial matches
  - closed-existing only transactions
  - multi-transaction leads
  - unmatched POS volume

Output:
  <output_dir>/leads_corrected.xlsx
  <output_dir>/pos_corrected.xlsx

The script is deterministic with ``--seed`` and keeps all rows scoped to the
requested warehouse.
"""

from __future__ import annotations

import argparse
import random
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
from faker import Faker
from openpyxl import load_workbook

from lead_match_runtime.business_rules import load_business_rules, get_fiscal_year

_RULES = load_business_rules()
DEFAULT_FISCAL_YEAR = get_fiscal_year(_RULES)
DEFAULT_NUM_LEADS = 400
DEFAULT_NUM_POS = 12000

WAREHOUSE_VOLUME_GUIDE = {
    569: 1651899,
    848: 1505473,
    823: 1459688,
    563: 1285728,
    827: 1127775,
    947: 1104727,
    652: 1083758,
    564: 1024805,
    115: 1001236,
    580: 933700,
    653: 854940,
    943: 765242,
    578: 741195,
    651: 615482,
    654: 604221,
    893: 587859,
    1581: 585995,
    767: 545547,
    579: 484278,
    729: 462664,
    1665: 457767,
    650: 420762,
    1661: 373740,
    1487: 357151,
    655: 297396,
    1663: 275664,
}

INDUSTRIES = [
    ("Retail", "Retail merchandise and local storefront operations"),
    ("Food Service", "Restaurants, cafes, catering, and food preparation"),
    ("Health", "Clinics, pharmacies, and wellness providers"),
    ("Construction", "Contractors, trades, and project services"),
    ("Professional Services", "Office-based business services and consulting"),
    ("Automotive", "Vehicle sales, repair, and parts operations"),
    ("Education", "Training, tutoring, and educational services"),
    ("Real Estate", "Property management and brokerage services"),
    ("Logistics", "Delivery, storage, and logistics providers"),
    ("Manufacturing", "Light manufacturing and fabrication"),
]

BUSINESS_SUFFIXES = ["LLC", "Inc", "Co", "Group", "Market", "Foods", "Supply", "Services", "Cafe", "Kitchen", "Partners"]
STREET_SUFFIX_VARIANTS = {
    "STREET": "ST",
    "AVENUE": "AVE",
    "ROAD": "RD",
    "DRIVE": "DR",
    "LANE": "LN",
    "BOULEVARD": "BLVD",
}

SCENARIO_WEIGHTS = {
    "exact_single": 0.10,
    "exact_multi": 0.08,
    "partial_single": 0.07,
    "partial_multi": 0.06,
    "closed_existing_only": 0.10,
    "late_cycle_unmatched": 0.05,
    "unmatched": 0.54,
}

PLAN_COLUMNS = [
    "warehouse_number",
    "num_leads",
    "num_pos",
    "seed",
    "fiscal_year",
    "exact_match_pct",
    "potential_match_pct",
    "closed_existing_pct",
]

DEFAULT_PLAN = {
    "warehouse_number": 569,
    "num_leads": DEFAULT_NUM_LEADS,
    "num_pos": DEFAULT_NUM_POS,
    "seed": 42,
    "fiscal_year": DEFAULT_FISCAL_YEAR,
    "exact_match_pct": 18.0,
    "potential_match_pct": 15.0,
    "closed_existing_pct": 10.0,
    "late_cycle_unmatched_pct": 5.0,
    "unmatched_pct": 52.0,
}


@dataclass
class LeadPlan:
    lead_index: int
    lead_id: str
    scenario: str
    canonical_name: str
    canonical_address: str
    city: str
    state: str
    zip_code: str
    email: str
    phone: str
    industry: str
    industry_description: str
    lead_fiscal_year: int
    lead_fiscal_period: int
    lead_week: int
    created_at: datetime
    match_count: int


def warehouse_default_counts(warehouse_number: int) -> tuple[int, int]:
    guide = WAREHOUSE_VOLUME_GUIDE.get(int(warehouse_number))
    if guide is None:
        return DEFAULT_NUM_LEADS, DEFAULT_NUM_POS
    if guide >= 1_500_000:
        return 300, 8000
    if guide >= 1_000_000:
        return 300, 8000
    if guide >= 500_000:
        return 300, 8000
    return 500, 10000


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate mock leads/POS Excel files.")
    parser.add_argument("--warehouse-number", type=int, default=None)
    parser.add_argument("--num-leads", type=int, default=None)
    parser.add_argument("--num-pos", type=int, default=None)
    parser.add_argument("--output-dir", default=None)
    parser.add_argument("--seed", type=int, default=None)
    parser.add_argument("--fiscal-year", type=int, default=None)
    parser.add_argument("--plan-csv", default=None, help="Optional CSV with one-row generation plan.")
    return parser.parse_args()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate mock leads/POS Excel files.")
    parser.add_argument("--warehouse-number", type=int, default=None)
    parser.add_argument("--num-leads", type=int, default=None)
    parser.add_argument("--num-pos", type=int, default=None)
    parser.add_argument("--output-dir", default=None)
    parser.add_argument("--seed", type=int, default=None)
    parser.add_argument("--fiscal-year", type=int, default=None)
    parser.add_argument("--plan-csv", default=None, help="Optional CSV with one-row generation plan.")
    return parser


def make_rng(seed: int) -> random.Random:
    return random.Random(seed)


def make_faker(seed: int) -> Faker:
    fake = Faker("en_US")
    Faker.seed(seed)
    fake.seed_instance(seed)
    return fake


def choose_industry(rng: random.Random) -> tuple[str, str]:
    return INDUSTRIES[rng.randrange(len(INDUSTRIES))]


def choose_fiscal_period(rng: random.Random, low: int, high: int) -> int:
    return rng.randint(low, high)


def period_to_timestamp(fiscal_year: int, fiscal_period: int, week: int, rng: random.Random) -> datetime:
    anchor = datetime(fiscal_year, 1, 1)
    days = (fiscal_period - 1) * 28 + (week - 1) * 7 + rng.randint(0, 3)
    hours = rng.randint(8, 18)
    minutes = rng.randint(0, 59)
    seconds = rng.randint(0, 59)
    return anchor + timedelta(days=days, hours=hours, minutes=minutes, seconds=seconds)


def phone_digits(rng: random.Random) -> str:
    area = rng.randint(200, 989)
    prefix = rng.randint(200, 989)
    line = rng.randint(1000, 9999)
    return f"{area:03d}-{prefix:03d}-{line:04d}"


def maybe_zip_plus4(zip_code: str, rng: random.Random) -> str:
    if rng.random() < 0.35:
        return f"{zip_code}-{rng.randint(1000, 9999)}"
    return zip_code


def load_plan_from_csv(path: str | None) -> dict[str, object]:
    if not path:
        return {}
    plan_path = Path(path)
    if not plan_path.exists():
        raise FileNotFoundError(f"Plan CSV not found: {plan_path}")
    df = pd.read_csv(plan_path)
    if df.empty:
        raise ValueError(f"Plan CSV is empty: {plan_path}")
    row = df.iloc[0].to_dict()
    return row


def normalize_plan_value(value, fallback):
    if pd.isna(value) or value == "":
        return fallback
    return value


def plan_to_weights(plan: dict[str, object] | None) -> dict[str, float]:
    plan = {**DEFAULT_PLAN, **(plan or {})}
    exact_pct = float(normalize_plan_value(plan.get("exact_match_pct"), 18.0))
    potential_pct = float(normalize_plan_value(plan.get("potential_match_pct"), 15.0))
    closed_existing_pct = float(normalize_plan_value(plan.get("closed_existing_pct"), 10.0))
    unmatched_pct = float(normalize_plan_value(plan.get("unmatched_pct"), 0.0))
    late_cycle_pct = float(normalize_plan_value(plan.get("late_cycle_unmatched_pct"), 5.0))

    exact_total = max(0.0, exact_pct)
    potential_total = max(0.0, potential_pct)
    closed_total = max(0.0, closed_existing_pct)
    late_total = max(0.0, late_cycle_pct)
    unmatched_total = max(0.0, unmatched_pct)

    remaining = 100.0 - (exact_total + potential_total + closed_total + late_total + unmatched_total)
    if remaining > 0:
        unmatched_total += remaining

    exact_single = exact_total * 0.55
    exact_multi = exact_total * 0.45
    partial_single = potential_total * 0.55
    partial_multi = potential_total * 0.45
    closed_existing_only = closed_total
    late_cycle_unmatched = late_total
    unmatched = unmatched_total

    total = exact_single + exact_multi + partial_single + partial_multi + closed_existing_only + late_cycle_unmatched + unmatched
    if total <= 0:
        return SCENARIO_WEIGHTS.copy()

    return {
        "exact_single": exact_single / total,
        "exact_multi": exact_multi / total,
        "partial_single": partial_single / total,
        "partial_multi": partial_multi / total,
        "closed_existing_only": closed_existing_only / total,
        "late_cycle_unmatched": late_cycle_unmatched / total,
        "unmatched": unmatched / total,
    }


def normalize_for_variant(text: str) -> str:
    text = re.sub(r"\s+", " ", text.strip())
    text = text.replace("&", "and")
    return text


def mutate_business_name(name: str, rng: random.Random) -> str:
    parts = normalize_for_variant(name).split()
    if not parts:
        return name

    mutations = {
        "NORTH": "N",
        "SOUTH": "S",
        "EAST": "E",
        "WEST": "W",
        "CENTER": "CTR",
        "CENTRE": "CTR",
        "COMPANY": "CO",
        "COMPANIES": "CO",
        "ASSOCIATES": "ASSOC",
        "SERVICES": "SVCS",
    }

    transformed = [mutations.get(part.upper(), part) for part in parts]
    if rng.random() < 0.5 and len(transformed) > 1:
        transformed.pop(rng.randrange(len(transformed)))
    if rng.random() < 0.4:
        transformed.append(rng.choice(["LLC", "Inc", "Co", "Group"]))
    return " ".join(transformed)


def mutate_address(address: str, rng: random.Random) -> str:
    parts = normalize_for_variant(address).upper().split()
    if not parts:
        return address

    if parts[-1] in STREET_SUFFIX_VARIANTS:
        parts[-1] = STREET_SUFFIX_VARIANTS[parts[-1]]
    elif len(parts) >= 2 and parts[-2] in STREET_SUFFIX_VARIANTS:
        parts[-2] = STREET_SUFFIX_VARIANTS[parts[-2]]

    if rng.random() < 0.35:
        parts.insert(-1, f"STE {rng.randint(100, 299)}")
    elif rng.random() < 0.35:
        parts[0] = str(int(re.sub(r"\D", "", parts[0]) or "1") + rng.randint(1, 9))
    return " ".join(parts)


def build_business_name(fake: Faker, rng: random.Random) -> str:
    left = fake.company().replace(",", "")
    if rng.random() < 0.45:
        return left
    adjective = rng.choice(
        ["North", "Pacific", "Summit", "Heritage", "Prime", "Evergreen", "Sunset", "Cascade", "Union", "Blue"]
    )
    noun = rng.choice(
        ["Market", "Kitchen", "Supply", "Foods", "Services", "Warehouse", "Dental", "Auto", "Studio", "Farm"]
    )
    suffix = rng.choice(BUSINESS_SUFFIXES)
    return f"{adjective} {noun} {suffix}"


def build_address(fake: Faker, rng: random.Random) -> tuple[str, str]:
    street = fake.street_address().replace(",", "")
    city = fake.city()
    state = fake.state_abbr().upper()
    zip_code = fake.zipcode()[:5].zfill(5)
    return street, f"{city}, {state} {zip_code}"


def build_family(fake: Faker, rng: random.Random, warehouse_number: int, fiscal_year: int, index: int) -> dict[str, object]:
    industry, description = choose_industry(rng)
    business_name = build_business_name(fake, rng)
    street, city_state_zip = build_address(fake, rng)
    city, state_zip = city_state_zip.split(",", 1)
    state, zip_code = state_zip.strip().split(" ")
    email_domain = fake.domain_name()
    email = f"{re.sub(r'[^a-z0-9]+', '.', business_name.lower()).strip('.')}@{email_domain}".replace("..", ".")
    phone = phone_digits(rng)
    return {
        "family_id": f"FAM-{fiscal_year}-{index + 1:04d}",
        "warehouse_number": str(warehouse_number),
        "business_name": business_name,
        "address_line_one": street,
        "city": city.strip(),
        "state": state.strip().upper(),
        "zip_code": zip_code.strip()[:5].zfill(5),
        "email": email,
        "phone": phone,
        "industry": industry,
        "industry_description": description,
        "lead_source": rng.choice(["ServiceNow", "Partner", "Web", "Referral", "Outbound"]),
        "type": rng.choice(["New Business", "Existing Account", "Expansion"]),
    }


def scenario_counts(num_leads: int, weights: dict[str, float]) -> dict[str, int]:
    counts = {name: int(round(num_leads * weight)) for name, weight in weights.items()}
    delta = num_leads - sum(counts.values())
    order = list(weights.keys())
    idx = 0
    while delta != 0:
        key = order[idx % len(order)]
        if delta > 0:
            counts[key] += 1
            delta -= 1
        elif counts[key] > 0:
            counts[key] -= 1
            delta += 1
        idx += 1
    return counts


def lead_scenario_sequence(num_leads: int, rng: random.Random, weights: dict[str, float]) -> list[str]:
    counts = scenario_counts(num_leads, weights)
    sequence: list[str] = []
    for scenario, count in counts.items():
        sequence.extend([scenario] * count)
    rng.shuffle(sequence)
    return sequence


def build_leads(num_leads: int, warehouse_number: int, fiscal_year: int, rng: random.Random, fake: Faker, weights: dict[str, float]) -> list[LeadPlan]:
    sequence = lead_scenario_sequence(num_leads, rng, weights)
    plans: list[LeadPlan] = []
    for idx, scenario in enumerate(sequence):
        family = build_family(fake, rng, warehouse_number, fiscal_year, idx)
        lead_period = choose_fiscal_period(rng, 2, 8 if scenario != "late_cycle_unmatched" else 13)
        lead_week = rng.randint(1, 5)
        created_at = period_to_timestamp(fiscal_year, lead_period, lead_week, rng)
        lead_id = f"LEAD-{created_at:%Y%m%d}-{idx + 1:05d}"
        match_count = 0
        if scenario in {"exact_single", "partial_single", "closed_existing_only"}:
            match_count = 1
        elif scenario in {"exact_multi", "partial_multi"}:
            match_count = rng.randint(2, 4)
        plans.append(
            LeadPlan(
                lead_index=idx,
                lead_id=lead_id,
                scenario=scenario,
                canonical_name=str(family["business_name"]),
                canonical_address=str(family["address_line_one"]),
                city=str(family["city"]),
                state=str(family["state"]),
                zip_code=str(family["zip_code"]),
                email=str(family["email"]),
                phone=str(family["phone"]),
                industry=str(family["industry"]),
                industry_description=str(family["industry_description"]),
                lead_fiscal_year=fiscal_year,
                lead_fiscal_period=lead_period,
                lead_week=lead_week,
                created_at=created_at,
                match_count=match_count,
            )
        )
    return plans


def generate_base_leads(
    num_leads: int,
    warehouse_number: int,
    fiscal_year: int,
    rng: random.Random,
    fake: Faker,
    weights: dict[str, float],
) -> list[LeadPlan]:
    return build_leads(num_leads, warehouse_number, fiscal_year, rng, fake, weights)


def lead_identity_row(plan: LeadPlan, rng: random.Random, warehouse_number: int) -> dict[str, object]:
    account_number = 10000000 + plan.lead_index + 1
    return {
        "lead_id": plan.lead_id,
        "warehouse_number": str(warehouse_number),
        "fiscal_year_lead": plan.lead_fiscal_year,
        "fiscal_period_lead": plan.lead_fiscal_period,
        "week": plan.lead_week,
        "updated_date": plan.created_at,
        "lead_source": rng.choice(["ServiceNow", "Partner", "Web", "Referral", "Outbound"]),
        "account_number": account_number,
        "membership_number": "",
        "lead_status": "Open",
        "confidence_level": "",
        "match_result": "",
        "type": rng.choice(["New Business", "Existing Account", "Expansion"]),
        "business_name": plan.canonical_name,
        "address_line_one": plan.canonical_address,
        "address_line_two": rng.choice(["", "Suite 100", "Bldg A", "Floor 2"]),
        "city": plan.city,
        "state": plan.state,
        "zip_code": plan.zip_code,
        "email": plan.email if rng.random() < 0.7 else "",
        "phone": plan.phone if rng.random() < 0.8 else "",
        "bd_industry": plan.industry,
        "industry_description": plan.industry_description,
        "customer_name": f"{plan.canonical_name} Account",
        "closed_fiscal_period": "",
        "closed_fiscal_year": "",
        "batch_id": f"BATCH-{warehouse_number}-{plan.created_at:%Y%m%d}",
        "load_date": plan.created_at,
        "updated_by": "mock_generator",
    }


def build_pos_row(
    plan: LeadPlan,
    scenario: str,
    warehouse_number: int,
    fiscal_year: int,
    rng: random.Random,
    fake: Faker,
    seq: int,
) -> dict[str, object]:
    same_family = scenario in {"exact_single", "exact_multi", "partial_single", "partial_multi", "closed_existing_only"}
    base_name = plan.canonical_name if same_family else build_business_name(fake, rng)
    base_address = plan.canonical_address if same_family else fake.street_address().replace(",", "")
    city = plan.city if same_family else fake.city()
    state = plan.state if same_family else fake.state_abbr().upper()
    zip_code = plan.zip_code if same_family else fake.zipcode()[:5].zfill(5)
    email = plan.email if same_family and rng.random() < 0.7 else f"{re.sub(r'[^a-z0-9]+', '.', (base_name or 'business').lower()).strip('.')}{rng.randint(1,999)}@{fake.domain_name()}"
    phone = plan.phone if same_family and rng.random() < 0.8 else phone_digits(rng)

    if scenario == "exact_single":
        business_name = base_name
        address_line_one = base_address
        pos_period = min(13, plan.lead_fiscal_period + rng.randint(0, 4))
        pos_year = fiscal_year
        pos_week = min(5, max(plan.lead_week, rng.randint(plan.lead_week, 5)))
    elif scenario == "exact_multi":
        business_name = base_name
        address_line_one = base_address
        pos_period = min(13, plan.lead_fiscal_period + seq)
        pos_year = fiscal_year
        pos_week = min(5, max(plan.lead_week, min(5, seq + plan.lead_week)))
    elif scenario == "partial_single":
        business_name = mutate_business_name(base_name, rng)
        address_line_one = mutate_address(base_address, rng)
        pos_period = min(13, plan.lead_fiscal_period + rng.randint(0, 3))
        pos_year = fiscal_year
        pos_week = min(5, max(plan.lead_week, rng.randint(plan.lead_week, 5)))
    elif scenario == "partial_multi":
        business_name = mutate_business_name(base_name, rng) if seq == 1 else base_name
        address_line_one = mutate_address(base_address, rng) if seq == 1 else base_address
        pos_period = min(13, plan.lead_fiscal_period + seq)
        pos_year = fiscal_year
        pos_week = min(5, max(plan.lead_week, min(5, seq + plan.lead_week)))
    elif scenario == "closed_existing_only":
        business_name = base_name
        address_line_one = base_address
        pos_year = fiscal_year
        pos_period = max(1, plan.lead_fiscal_period - rng.randint(1, 4))
        pos_week = max(1, plan.lead_week - rng.randint(0, 2))
    else:
        business_name = base_name if rng.random() < 0.15 else mutate_business_name(base_name, rng)
        address_line_one = base_address if rng.random() < 0.15 else mutate_address(base_address, rng)
        pos_year = fiscal_year
        pos_period = rng.randint(1, 13)
        pos_week = rng.randint(1, 5)

    pos_date = period_to_timestamp(pos_year, pos_period, pos_week, rng)
    order_amount = make_order_amount(rng, scenario)
    first_name = fake.first_name()
    last_name = fake.last_name()
    membership_number = f"{rng.randint(1000000000, 9999999999)}"
    account_number = 10000000 + plan.lead_index + 1
    sales_reference_id = f"SR-{warehouse_number}-{pos_date:%Y%m%d}-{seq:03d}"
    address_line_two = rng.choice(["", "Unit 2", "Apt 5", "Suite 210", "Dock 4"])
    shop_type = rng.choices(["Delivery", "In-Warehouse", "Pickup"], weights=[0.3, 0.5, 0.2], k=1)[0]
    bd_industry, industry_description = choose_industry(rng)

    return {
        "pos_id": f"POS-{pos_date:%Y%m%d}-{plan.lead_index + 1:05d}-{seq:02d}",
        "warehouse_number": str(warehouse_number),
        "fiscal_year_transaction": pos_year,
        "fiscal_period_transaction": pos_period,
        "week": pos_week,
        "account_number": account_number,
        "membership_number": membership_number,
        "sales_reference_id": sales_reference_id,
        "business_name": business_name,
        "first_name": first_name,
        "last_name": last_name,
        "address_line_one": address_line_one,
        "address_line_two": address_line_two,
        "city": city,
        "state": state,
        "zip_code": maybe_zip_plus4(zip_code, rng),
        "email": email,
        "phone": phone,
        "order_amount": order_amount,
        "shop_type": shop_type,
        "bd_industry": bd_industry,
        "industry_description": industry_description,
        "load_date": pos_date,
        "updated_date": pos_date,
        "expected_relation": scenario,
        "expected_lead_id": plan.lead_id,
    }


def inject_exact_matches(
    plans: list[LeadPlan],
    warehouse_number: int,
    fiscal_year: int,
    rng: random.Random,
    fake: Faker,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for plan in plans:
        if plan.scenario.startswith("exact"):
            for seq in range(1, plan.match_count + 1):
                rows.append(build_pos_row(plan, plan.scenario, warehouse_number, fiscal_year, rng, fake, seq))
    return rows


def inject_partial_matches(
    plans: list[LeadPlan],
    warehouse_number: int,
    fiscal_year: int,
    rng: random.Random,
    fake: Faker,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for plan in plans:
        if plan.scenario.startswith("partial"):
            for seq in range(1, plan.match_count + 1):
                rows.append(build_pos_row(plan, plan.scenario, warehouse_number, fiscal_year, rng, fake, seq))
    return rows


def inject_closed_existing_cases(
    plans: list[LeadPlan],
    warehouse_number: int,
    fiscal_year: int,
    rng: random.Random,
    fake: Faker,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for plan in plans:
        if plan.scenario == "closed_existing_only":
            rows.append(build_pos_row(plan, plan.scenario, warehouse_number, fiscal_year, rng, fake, 1))
    return rows


def inject_multi_transaction_cases(
    plans: list[LeadPlan],
    warehouse_number: int,
    fiscal_year: int,
    rng: random.Random,
    fake: Faker,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for plan in plans:
        if plan.scenario in {"exact_multi", "partial_multi"}:
            for seq in range(1, plan.match_count + 1):
                rows.append(build_pos_row(plan, plan.scenario, warehouse_number, fiscal_year, rng, fake, seq))
    return rows


def generate_noise_pos(
    plans: list[LeadPlan],
    num_pos: int,
    warehouse_number: int,
    fiscal_year: int,
    rng: random.Random,
    fake: Faker,
) -> list[dict[str, object]]:
    pos_rows, _ = generate_pos_rows(plans, num_pos, warehouse_number, fiscal_year, rng, fake)
    return pos_rows


def make_order_amount(rng: random.Random, scenario: str) -> float:
    if scenario in {"exact_single", "exact_multi", "partial_single", "partial_multi"}:
        bands = [(0.55, (99, 499)), (0.25, (500, 1499)), (0.20, (1500, 4500))]
    elif scenario == "closed_existing_only":
        bands = [(0.45, (49, 299)), (0.35, (300, 999)), (0.20, (1000, 2400))]
    else:
        bands = [(0.70, (19, 299)), (0.20, (300, 899)), (0.10, (900, 2500))]
    roll = rng.random()
    cumulative = 0.0
    for weight, (low, high) in bands:
        cumulative += weight
        if roll <= cumulative:
            return round(rng.uniform(low, high), 2)
    low, high = bands[-1][1]
    return round(rng.uniform(low, high), 2)


def generate_pos_rows(
    plans: list[LeadPlan],
    num_pos: int,
    warehouse_number: int,
    fiscal_year: int,
    rng: random.Random,
    fake: Faker,
) -> tuple[list[dict[str, object]], dict[str, int]]:
    pos_rows: list[dict[str, object]] = []
    counts = {"linked": 0, "unmatched": 0, "exact": 0, "partial": 0, "closed_existing": 0}

    for plan in plans:
        if plan.scenario == "unmatched" or plan.match_count == 0:
            continue
        for seq in range(1, plan.match_count + 1):
            row = build_pos_row(plan, plan.scenario, warehouse_number, fiscal_year, rng, fake, seq)
            pos_rows.append(row)
            counts["linked"] += 1
            if plan.scenario.startswith("exact"):
                counts["exact"] += 1
            elif plan.scenario.startswith("partial"):
                counts["partial"] += 1
            elif plan.scenario == "closed_existing_only":
                counts["closed_existing"] += 1

    remaining = max(0, num_pos - len(pos_rows))
    for seq in range(1, remaining + 1):
        fake_plan = rng.choice(plans)
        scenario = "unmatched"
        row = build_pos_row(fake_plan, scenario, warehouse_number, fiscal_year, rng, fake, seq)
        row["expected_lead_id"] = ""
        pos_rows.append(row)
        counts["unmatched"] += 1

    rng.shuffle(pos_rows)
    return pos_rows[:num_pos], counts


def add_lead_specific_variants(leads: list[dict[str, object]]) -> None:
    # No-op placeholder for future enrichment. Kept to make the generator easy to extend.
    return


def format_excel_columns(path: Path, text_columns: Iterable[str], datetime_columns: Iterable[str]) -> None:
    wb = load_workbook(path)
    ws = wb.active
    header_map = {ws.cell(row=1, column=col_idx).value: col_idx for col_idx in range(1, ws.max_column + 1)}

    for column in text_columns:
        if column not in header_map:
            continue
        col_idx = header_map[column]
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = "@"

    for column in datetime_columns:
        if column not in header_map:
            continue
        col_idx = header_map[column]
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"

    wb.save(path)


def write_excel_files(
    leads_df: pd.DataFrame,
    pos_df: pd.DataFrame,
    output_dir: Path,
    lead_text_cols: list[str],
    lead_datetime_cols: list[str],
    pos_text_cols: list[str],
    pos_datetime_cols: list[str],
) -> tuple[Path, Path]:
    leads_path = output_dir / "leads_corrected.xlsx"
    pos_path = output_dir / "pos_corrected.xlsx"
    write_workbook(leads_df, leads_path, lead_text_cols, lead_datetime_cols)
    write_workbook(pos_df, pos_path, pos_text_cols, pos_datetime_cols)
    return leads_path, pos_path


def write_workbook(df: pd.DataFrame, output_path: Path, text_columns: list[str], datetime_columns: list[str]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False, engine="openpyxl")
    format_excel_columns(output_path, text_columns=text_columns, datetime_columns=datetime_columns)


def print_summary(warehouse_number: int, plans: list[LeadPlan], pos_counts: dict[str, int], num_pos: int, weights: dict[str, float], leads_path: Path, pos_path: Path) -> None:
    scenario_totals: dict[str, int] = {}
    for plan in plans:
        scenario_totals[plan.scenario] = scenario_totals.get(plan.scenario, 0) + 1

    guide = WAREHOUSE_VOLUME_GUIDE.get(warehouse_number)
    print(f"Generated mock data for warehouse {warehouse_number}")
    if guide is not None:
        print(f"Reference warehouse volume guide: {guide:,}")
    print(f"Leads: {len(plans):,}")
    print(f"POS target: {num_pos:,}")
    print("Configured mix:")
    for key, value in weights.items():
        print(f"  - {key}: {value:.2%}")
    print("Scenario counts:")
    for key in ["exact_single", "exact_multi", "partial_single", "partial_multi", "closed_existing_only", "late_cycle_unmatched", "unmatched"]:
        print(f"  - {key}: {scenario_totals.get(key, 0):,}")
    print("POS build summary:")
    print(f"  - linked rows: {pos_counts['linked']:,}")
    print(f"  - exact-linked rows: {pos_counts['exact']:,}")
    print(f"  - partial-linked rows: {pos_counts['partial']:,}")
    print(f"  - closed-existing rows: {pos_counts['closed_existing']:,}")
    print(f"  - unmatched rows: {pos_counts['unmatched']:,}")
    print("Output files:")
    print(f"  - {leads_path}")
    print(f"  - {pos_path}")


def main() -> int:
    args = parse_args()
    plan = load_plan_from_csv(args.plan_csv)
    warehouse_number = int(normalize_plan_value(args.warehouse_number, normalize_plan_value(plan.get("warehouse_number"), DEFAULT_PLAN["warehouse_number"])))
    default_leads, default_pos = warehouse_default_counts(warehouse_number)
    num_leads = int(normalize_plan_value(args.num_leads, normalize_plan_value(plan.get("num_leads"), default_leads)))
    num_pos = int(normalize_plan_value(args.num_pos, normalize_plan_value(plan.get("num_pos"), default_pos)))
    fiscal_year = int(normalize_plan_value(args.fiscal_year, normalize_plan_value(plan.get("fiscal_year"), DEFAULT_PLAN["fiscal_year"])))
    seed = int(normalize_plan_value(args.seed, normalize_plan_value(plan.get("seed"), DEFAULT_PLAN["seed"])))
    rng = make_rng(seed)
    fake = make_faker(seed)
    weights = plan_to_weights(plan)

    output_dir = Path(args.output_dir or Path.cwd() / str(warehouse_number))
    output_dir.mkdir(parents=True, exist_ok=True)

    plans = generate_base_leads(num_leads, warehouse_number, fiscal_year, rng, fake, weights)
    lead_rows = [lead_identity_row(plan, rng, warehouse_number) for plan in plans]
    add_lead_specific_variants(lead_rows)
    pos_rows, pos_counts = generate_pos_rows(plans, num_pos, warehouse_number, fiscal_year, rng, fake)

    leads_df = pd.DataFrame(lead_rows)
    pos_df = pd.DataFrame(pos_rows)

    # Keep workbook shapes friendly for downstream preprocessors.
    lead_text_cols = [
        "lead_id",
        "warehouse_number",
        "lead_source",
        "account_number",
        "membership_number",
        "lead_status",
        "confidence_level",
        "match_result",
        "type",
        "business_name",
        "address_line_one",
        "address_line_two",
        "city",
        "state",
        "zip_code",
        "email",
        "phone",
        "bd_industry",
        "industry_description",
        "customer_name",
        "batch_id",
        "updated_by",
    ]
    lead_datetime_cols = ["updated_date", "load_date"]

    pos_text_cols = [
        "pos_id",
        "warehouse_number",
        "account_number",
        "membership_number",
        "sales_reference_id",
        "business_name",
        "first_name",
        "last_name",
        "address_line_one",
        "address_line_two",
        "city",
        "state",
        "zip_code",
        "email",
        "phone",
        "shop_type",
        "bd_industry",
        "industry_description",
        "expected_relation",
        "expected_lead_id",
    ]
    pos_datetime_cols = ["load_date", "updated_date"]

    leads_path = output_dir / "leads_corrected.xlsx"
    pos_path = output_dir / "pos_corrected.xlsx"
    write_excel_files(leads_df, pos_df, output_dir, lead_text_cols, lead_datetime_cols, pos_text_cols, pos_datetime_cols)

    print_summary(warehouse_number, plans, pos_counts, num_pos, weights, leads_path, pos_path)
    print(f"Workbook directory: {output_dir}")
    print(f"Lead workbook rows: {len(leads_df):,}")
    print(f"POS workbook rows: {len(pos_df):,}")
    print("Tip: point the matching pipeline at this directory when you test warehouse-scoped matching.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
