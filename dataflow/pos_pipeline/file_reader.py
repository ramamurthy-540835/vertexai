"""
Multi-format file reader.

Provides two APIs:

  read_file_to_dicts(content, filename)
      Loads the entire file into memory and returns a list of dicts.
      Kept for backwards compatibility — use only for small files.

  iter_file_chunks(blob, filename, chunk_size)
      STREAMING generator. Reads from a GCS blob and yields chunks of
      `chunk_size` rows at a time. Never holds more than one chunk in memory.
      Use this for large files (CSV > 100MB, Excel > 50MB).

Returns dicts where keys are the source column headers (normalized: stripped
and lowercased) and values are Python primitives. Mapping to DB columns
happens later in schema_utils.apply_field_map.
"""

import csv
import io
import json
import logging
import os
from typing import List, Dict, Any, Iterator

logger = logging.getLogger(__name__)


SUPPORTED_EXTENSIONS = {
    ".csv", ".tsv", ".psv", ".txt",
    ".xlsx", ".xls",
    ".json", ".jsonl",
    ".parquet",
}


# ─────────────────────────────────────────────────────────────
# Public API — STREAMING (preferred for large files)
# ─────────────────────────────────────────────────────────────

def iter_file_chunks(
    blob,
    filename: str,
    chunk_size: int = 10000,
) -> Iterator[List[Dict[str, Any]]]:
    """
    Stream a GCS blob and yield chunks of `chunk_size` row dicts.

    Memory-safe for large files: never holds more than one chunk in memory.

    Args:
        blob: google.cloud.storage.Blob object (not bytes)
        filename: filename used for extension detection
        chunk_size: max rows per yielded chunk

    Yields:
        list[dict] — one chunk of rows at a time
    """
    ext = os.path.splitext(filename.lower())[1]
    logger.info(f"Streaming file {filename} (ext={ext}, chunk_size={chunk_size})")

    if ext in (".csv", ".txt"):
        yield from _iter_delimited(blob, filename, delimiter=",", chunk_size=chunk_size)
    elif ext == ".tsv":
        yield from _iter_delimited(blob, filename, delimiter="\t", chunk_size=chunk_size)
    elif ext == ".psv":
        yield from _iter_delimited(blob, filename, delimiter="|", chunk_size=chunk_size)
    elif ext in (".xlsx", ".xls"):
        yield from _iter_excel(blob, chunk_size=chunk_size)
    elif ext in (".json", ".jsonl", ".parquet"):
        # These formats are not commonly large in this pipeline.
        # Fall back to full-load then chunk.
        content = blob.download_as_bytes()
        rows = read_file_to_dicts(content, filename)
        for i in range(0, len(rows), chunk_size):
            yield rows[i:i + chunk_size]
    else:
        # Unknown extension — try CSV streaming
        logger.warning(f"Unknown extension {ext}, attempting CSV stream")
        yield from _iter_delimited(blob, filename, delimiter=",", chunk_size=chunk_size)


def _iter_delimited(
    blob,
    filename: str,
    delimiter: str,
    chunk_size: int,
) -> Iterator[List[Dict[str, Any]]]:
    """
    Stream CSV/TSV/PSV from GCS line-by-line and yield row chunks.
    Tries multiple encodings to handle BOM / Windows files.
    """
    last_err = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            # Open a streaming connection to GCS — does NOT download whole file
            with blob.open("rb") as raw_stream:
                text_stream = io.TextIOWrapper(raw_stream, encoding=enc, newline="")
                reader = csv.reader(text_stream, delimiter=delimiter)

                # Read header row
                try:
                    raw_headers = next(reader)
                except StopIteration:
                    logger.warning(f"{filename} is empty (no header row)")
                    return

                headers = [
                    str(h).strip().lower() if h is not None else ""
                    for h in raw_headers
                ]
                logger.info(
                    f"Streaming CSV headers ({len(headers)}): {headers}"
                )

                chunk: List[Dict[str, Any]] = []
                total_rows = 0

                for raw_row in reader:
                    # Skip blank rows
                    if not raw_row or all(
                        (v is None or str(v).strip() == "") for v in raw_row
                    ):
                        continue

                    row_dict = {
                        headers[i]: raw_row[i]
                        for i in range(min(len(headers), len(raw_row)))
                        if headers[i]
                    }
                    chunk.append(row_dict)
                    total_rows += 1

                    if len(chunk) >= chunk_size:
                        logger.info(
                            f"Yielding chunk of {len(chunk)} rows "
                            f"(total so far: {total_rows})"
                        )
                        yield chunk
                        chunk = []

                # Yield remaining tail
                if chunk:
                    logger.info(
                        f"Yielding final chunk of {len(chunk)} rows "
                        f"(total: {total_rows})"
                    )
                    yield chunk

                logger.info(
                    f"Streaming done: {total_rows} rows from {filename} "
                    f"(encoding={enc})"
                )
                return

        except UnicodeDecodeError as e:
            last_err = e
            logger.info(f"Encoding {enc} failed, trying next…")
            continue

    raise ValueError(
        f"Could not decode {filename} with any supported encoding "
        f"(last error: {last_err})"
    )


def _iter_excel(blob, chunk_size: int) -> Iterator[List[Dict[str, Any]]]:
    """
    Stream Excel (.xlsx / .xls) rows in chunks.

    Note: Excel files cannot be streamed directly from GCS — the entire
    file is downloaded (compressed Excel files are typically much smaller
    than the equivalent CSV). However openpyxl's read_only mode streams
    rows from the unzipped file without loading all cells into memory.
    """
    content = blob.download_as_bytes()

    # Try xlsx first
    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        raw_headers = next(rows_iter, None)
        if not raw_headers:
            logger.warning("Excel file has no header row")
            return

        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in raw_headers
        ]
        logger.info(f"Streaming Excel headers ({len(headers)}): {headers}")

        chunk: List[Dict[str, Any]] = []
        total_rows = 0

        for row in rows_iter:
            if all(v is None or v == "" for v in row):
                continue  # skip blank rows
            row_dict = {
                headers[i]: (row[i] if i < len(row) else None)
                for i in range(len(headers))
                if headers[i]
            }
            chunk.append(row_dict)
            total_rows += 1

            if len(chunk) >= chunk_size:
                logger.info(
                    f"Yielding Excel chunk of {len(chunk)} rows "
                    f"(total so far: {total_rows})"
                )
                yield chunk
                chunk = []

        if chunk:
            logger.info(
                f"Yielding final Excel chunk of {len(chunk)} rows "
                f"(total: {total_rows})"
            )
            yield chunk

        wb.close()
        logger.info(f"Excel streaming done: {total_rows} rows")
        return

    except Exception as e:
        logger.info(f"openpyxl failed ({e}); trying xlrd fallback for .xls")

    # Fall back to xlrd for old .xls — does not support streaming, but
    # .xls files are typically smaller (format limited to 65k rows per sheet)
    import xlrd
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    headers = [str(ws.cell_value(0, c)).strip().lower() for c in range(ws.ncols)]

    chunk: List[Dict[str, Any]] = []
    total_rows = 0
    for r in range(1, ws.nrows):
        row_dict = {
            headers[c]: ws.cell_value(r, c)
            for c in range(ws.ncols)
            if headers[c]
        }
        if any(v not in (None, "") for v in row_dict.values()):
            chunk.append(row_dict)
            total_rows += 1
            if len(chunk) >= chunk_size:
                yield chunk
                chunk = []
    if chunk:
        yield chunk

    logger.info(f"xls read: {total_rows} rows")


# ─────────────────────────────────────────────────────────────
# Public API — FULL LOAD (kept for backwards compatibility)
# ─────────────────────────────────────────────────────────────

def read_file_to_dicts(content: bytes, filename: str) -> List[Dict[str, Any]]:
    """
    Dispatch to the right reader based on file extension.
    Returns a list of row dicts. Header names are normalized.

    ⚠️  Loads entire file into memory. For large files use iter_file_chunks().
    """

    if not content:
        logger.warning(f"Skipping zero-byte/corrupted file: {filename}")
        return []
    ext = os.path.splitext(filename.lower())[1]
    logger.info(f"Reading file {filename} with extension {ext}")

    if ext in (".csv", ".txt"):
        return _read_delimited(content, delimiter=",")
    if ext == ".tsv":
        return _read_delimited(content, delimiter="\t")
    if ext == ".psv":
        return _read_delimited(content, delimiter="|")
    if ext in (".xlsx", ".xls"):
        return _read_excel(content)
    if ext == ".json":
        return _read_json(content)
    if ext == ".jsonl":
        return _read_jsonl(content)
    if ext == ".parquet":
        return _read_parquet(content)

    logger.warning(f"Unknown extension {ext}, attempting CSV parse")
    return _read_delimited(content, delimiter=",")


def _read_delimited(content: bytes, delimiter: str = ",") -> List[Dict[str, Any]]:
    """Full-load CSV/TSV/PSV reader. Use iter_file_chunks for large files."""
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = content.decode(enc)
            reader = csv.reader(io.StringIO(text), delimiter=delimiter)

            try:
                raw_headers = next(reader)
            except StopIteration:
                logger.warning("CSV file is empty (no header row)")
                return []

            headers = [
                str(h).strip().lower() if h is not None else ""
                for h in raw_headers
            ]
            logger.info(f"CSV normalized headers ({len(headers)}): {headers}")

            rows: List[Dict[str, Any]] = []
            for raw_row in reader:
                if not raw_row or all(
                    (v is None or str(v).strip() == "") for v in raw_row
                ):
                    continue
                row_dict = {
                    headers[i]: raw_row[i]
                    for i in range(min(len(headers), len(raw_row)))
                    if headers[i]
                }
                rows.append(row_dict)

            logger.info(
                f"Delimited read: {len(rows)} rows "
                f"(encoding={enc}, headers={len([h for h in headers if h])})"
            )
            return rows
        except UnicodeDecodeError:
            continue

    raise ValueError("Could not decode file with any supported encoding")


def _read_excel(content: bytes) -> List[Dict[str, Any]]:
    """Full-load Excel reader. Use iter_file_chunks for large files."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return []
        headers = [str(h).strip().lower() if h is not None else "" for h in headers]

        out: List[Dict[str, Any]] = []
        for row in rows_iter:
            if all(v is None or v == "" for v in row):
                continue
            out.append({
                headers[i]: (row[i] if i < len(row) else None)
                for i in range(len(headers))
                if headers[i]
            })
        logger.info(f"Excel (xlsx) read: {len(out)} rows")
        return out
    except Exception as e:
        logger.info(f"openpyxl failed ({e}); trying xlrd fallback for .xls")

    import xlrd
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    headers = [str(ws.cell_value(0, c)).strip().lower() for c in range(ws.ncols)]
    out = []
    for r in range(1, ws.nrows):
        row_dict = {
            headers[c]: ws.cell_value(r, c)
            for c in range(ws.ncols)
            if headers[c]
        }
        if any(v not in (None, "") for v in row_dict.values()):
            out.append(row_dict)
    logger.info(f"Excel (xls) read: {len(out)} rows")
    return out


def _read_json(content: bytes) -> List[Dict[str, Any]]:
    """Parse JSON as list-of-objects or {rows: [...]} envelope."""
    text = content.decode("utf-8-sig")
    data = json.loads(text)
    if isinstance(data, list):
        rows = [r for r in data if isinstance(r, dict)]
    elif isinstance(data, dict) and isinstance(data.get("rows"), list):
        rows = [r for r in data["rows"] if isinstance(r, dict)]
    else:
        raise ValueError("JSON must be a list of objects, or {rows: [...]}")
    rows = [{k.strip().lower(): v for k, v in r.items()} for r in rows]
    logger.info(f"JSON read: {len(rows)} rows")
    return rows


def _read_jsonl(content: bytes) -> List[Dict[str, Any]]:
    """Parse newline-delimited JSON."""
    text = content.decode("utf-8-sig")
    rows = [json.loads(line) for line in text.splitlines() if line.strip()]
    rows = [{k.strip().lower(): v for k, v in r.items()} for r in rows]
    logger.info(f"JSONL read: {len(rows)} rows")
    return rows


def _read_parquet(content: bytes) -> List[Dict[str, Any]]:
    """Parse Parquet to list of row dicts."""
    import pyarrow.parquet as pq
    table = pq.read_table(io.BytesIO(content))
    rows = table.to_pylist()
    rows = [{k.strip().lower(): v for k, v in r.items()} for r in rows]
    logger.info(f"Parquet read: {len(rows)} rows")
    return rows